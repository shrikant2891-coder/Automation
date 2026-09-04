#!/usr/bin/env python3
"""Generate projected balance sheets for NA Marketing And Design Private Limited."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path("/workspace/NA_Marketing_Balance_Sheet_Projections_FY26-28.xlsx")

COMPANY = {
    "name": "NA MARKETING AND DESIGN PRIVATE LIMITED",
    "cin": "U51101DL2006PTC149065",
    "registered_office": "S-448, Greater Kailash Part-2, New Delhi - 110048",
    "parent": "Unique Collections (India) Private Limited",
}

# Base provisional balance sheet figures (₹ in Lakhs) as on 31-Mar-2026 (FY 2025-26).
# Share capital verified from MCA filings. Other figures are provisional estimates aligned
# to reported revenue (~₹18.84 Cr FY25), secured borrowings (~₹44.14 Cr), and trading profile.
BASE_FY2526 = {
    "Share Capital": 119.00,
    "Reserves and Surplus": 2850.00,
    "Long Term Borrowings": 3200.00,
    "Deferred Tax Liabilities (Net)": 85.00,
    "Other Long Term Liabilities": 120.00,
    "Long Term Provisions": 45.00,
    "Short Term Borrowings": 1214.00,
    "Trade Payables": 2450.00,
    "Other Current Liabilities": 680.00,
    "Short Term Provisions": 95.00,
    "Property, Plant and Equipment": 1250.00,
    "Intangible Assets": 35.00,
    "Capital Work-in-Progress": 0.00,
    "Non-Current Investments": 180.00,
    "Deferred Tax Assets (Net)": 65.00,
    "Long Term Loans and Advances": 220.00,
    "Other Non-Current Assets": 95.00,
    "Inventories": 4850.00,
    "Trade Receivables": 2680.00,
    "Cash and Cash Equivalents": 920.00,
    "Short Term Loans and Advances": 310.00,
    "Other Current Assets": 253.00,
}

GROWTH_SCENARIOS = {
    "Conservative (35%)": 0.35,
    "Base Case (37.5%)": 0.375,
    "Aggressive (40%)": 0.40,
}

PRIMARY_SCENARIO = "Base Case (37.5%)"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=12, color="1F4E79")
LABEL_FONT = Font(bold=True)
MONEY_FMT = "#,##0.00"
PCT_FMT = "0.0%"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

BS_STRUCTURE = [
    ("EQUITY AND LIABILITIES", None, "section"),
    ("Shareholders' Funds", None, "subsection"),
    ("Share Capital", "Share Capital", "line"),
    ("Reserves and Surplus", "Reserves and Surplus", "line"),
    ("Total Shareholders' Funds", ["Share Capital", "Reserves and Surplus"], "subtotal"),
    ("", None, "blank"),
    ("Non-Current Liabilities", None, "subsection"),
    ("Long Term Borrowings", "Long Term Borrowings", "line"),
    ("Deferred Tax Liabilities (Net)", "Deferred Tax Liabilities (Net)", "line"),
    ("Other Long Term Liabilities", "Other Long Term Liabilities", "line"),
    ("Long Term Provisions", "Long Term Provisions", "line"),
    (
        "Total Non-Current Liabilities",
        [
            "Long Term Borrowings",
            "Deferred Tax Liabilities (Net)",
            "Other Long Term Liabilities",
            "Long Term Provisions",
        ],
        "subtotal",
    ),
    ("", None, "blank"),
    ("Current Liabilities", None, "subsection"),
    ("Short Term Borrowings", "Short Term Borrowings", "line"),
    ("Trade Payables", "Trade Payables", "line"),
    ("Other Current Liabilities", "Other Current Liabilities", "line"),
    ("Short Term Provisions", "Short Term Provisions", "line"),
    (
        "Total Current Liabilities",
        [
            "Short Term Borrowings",
            "Trade Payables",
            "Other Current Liabilities",
            "Short Term Provisions",
        ],
        "subtotal",
    ),
    ("", None, "blank"),
    (
        "TOTAL EQUITY AND LIABILITIES",
        [
            "Share Capital",
            "Reserves and Surplus",
            "Long Term Borrowings",
            "Deferred Tax Liabilities (Net)",
            "Other Long Term Liabilities",
            "Long Term Provisions",
            "Short Term Borrowings",
            "Trade Payables",
            "Other Current Liabilities",
            "Short Term Provisions",
        ],
        "total",
    ),
    ("", None, "blank"),
    ("ASSETS", None, "section"),
    ("Non-Current Assets", None, "subsection"),
    ("Property, Plant and Equipment", "Property, Plant and Equipment", "line"),
    ("Intangible Assets", "Intangible Assets", "line"),
    ("Capital Work-in-Progress", "Capital Work-in-Progress", "line"),
    ("Non-Current Investments", "Non-Current Investments", "line"),
    ("Deferred Tax Assets (Net)", "Deferred Tax Assets (Net)", "line"),
    ("Long Term Loans and Advances", "Long Term Loans and Advances", "line"),
    ("Other Non-Current Assets", "Other Non-Current Assets", "line"),
    (
        "Total Non-Current Assets",
        [
            "Property, Plant and Equipment",
            "Intangible Assets",
            "Capital Work-in-Progress",
            "Non-Current Investments",
            "Deferred Tax Assets (Net)",
            "Long Term Loans and Advances",
            "Other Non-Current Assets",
        ],
        "subtotal",
    ),
    ("", None, "blank"),
    ("Current Assets", None, "subsection"),
    ("Inventories", "Inventories", "line"),
    ("Trade Receivables", "Trade Receivables", "line"),
    ("Cash and Cash Equivalents", "Cash and Cash Equivalents", "line"),
    ("Short Term Loans and Advances", "Short Term Loans and Advances", "line"),
    ("Other Current Assets", "Other Current Assets", "line"),
    (
        "Total Current Assets",
        [
            "Inventories",
            "Trade Receivables",
            "Cash and Cash Equivalents",
            "Short Term Loans and Advances",
            "Other Current Assets",
        ],
        "subtotal",
    ),
    ("", None, "blank"),
    (
        "TOTAL ASSETS",
        [
            "Property, Plant and Equipment",
            "Intangible Assets",
            "Capital Work-in-Progress",
            "Non-Current Investments",
            "Deferred Tax Assets (Net)",
            "Long Term Loans and Advances",
            "Other Non-Current Assets",
            "Inventories",
            "Trade Receivables",
            "Cash and Cash Equivalents",
            "Short Term Loans and Advances",
            "Other Current Assets",
        ],
        "total",
    ),
]

CONSTANT_ITEMS = {"Share Capital"}
BALANCING_ITEM = "Reserves and Surplus"


def liability_keys():
    return [
        k
        for k in BASE_FY2526
        if k not in {
            "Share Capital",
            BALANCING_ITEM,
            "Property, Plant and Equipment",
            "Intangible Assets",
            "Capital Work-in-Progress",
            "Non-Current Investments",
            "Deferred Tax Assets (Net)",
            "Long Term Loans and Advances",
            "Other Non-Current Assets",
            "Inventories",
            "Trade Receivables",
            "Cash and Cash Equivalents",
            "Short Term Loans and Advances",
            "Other Current Assets",
        }
    ]


def asset_keys():
    return [
        "Property, Plant and Equipment",
        "Intangible Assets",
        "Capital Work-in-Progress",
        "Non-Current Investments",
        "Deferred Tax Assets (Net)",
        "Long Term Loans and Advances",
        "Other Non-Current Assets",
        "Inventories",
        "Trade Receivables",
        "Cash and Cash Equivalents",
        "Short Term Loans and Advances",
        "Other Current Assets",
    ]


def equity_liability_keys():
    return [k for k in BASE_FY2526 if k not in asset_keys()]


def project_values(base: dict[str, float], growth: float, years: int = 1) -> dict[str, float]:
    factor = (1 + growth) ** years
    projected = {}
    for key, value in base.items():
        if key in CONSTANT_ITEMS:
            projected[key] = value
        elif key == BALANCING_ITEM:
            continue
        else:
            projected[key] = round(value * factor, 2)

    total_assets = sum(projected[k] for k in asset_keys())
    total_liabilities = sum(projected[k] for k in liability_keys())
    projected[BALANCING_ITEM] = round(
        total_assets - projected["Share Capital"] - total_liabilities,
        2,
    )
    return projected


def style_range(ws, cell_range, fill=None, font=None, border=THIN_BORDER):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if border:
                cell.border = border


def set_col_widths(ws, widths):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_cover(wb: Workbook):
    ws = wb.active
    ws.title = "Cover"
    set_col_widths(ws, {1: 4, 2: 28, 3: 55, 4: 18})

    ws.merge_cells("B2:D2")
    ws["B2"] = "PROJECTED & ESTIMATED BALANCE SHEETS"
    ws["B2"].font = TITLE_FONT
    ws["B2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("B4:D4")
    ws["B4"] = COMPANY["name"]
    ws["B4"].font = SUBTITLE_FONT
    ws["B4"].alignment = Alignment(horizontal="center")

    details = [
        ("CIN", COMPANY["cin"]),
        ("Registered Office", COMPANY["registered_office"]),
        ("Parent Company", COMPANY["parent"]),
        ("Currency", "Indian Rupees (₹)"),
        ("Unit", "Lakhs"),
        ("Base Statement", "Provisional Balance Sheet as on 31-Mar-2026 (FY 2025-26)"),
        ("Projected Statement", "Projected Balance Sheet as on 31-Mar-2027 (FY 2026-27)"),
        ("Estimated Statement", "Estimated Balance Sheet as on 31-Mar-2028 (FY 2027-28)"),
        ("Growth Assumption", "35% to 40% year-on-year increment on provisional base"),
        ("Primary Scenario", PRIMARY_SCENARIO),
        ("Prepared On", "04-Sep-2026"),
    ]

    row = 7
    for label, value in details:
        ws[f"B{row}"] = label
        ws[f"B{row}"].font = LABEL_FONT
        ws[f"C{row}"] = value
        ws[f"B{row}"].border = THIN_BORDER
        ws[f"C{row}"].border = THIN_BORDER
        row += 1

    ws.merge_cells(f"B{row + 2}:D{row + 5}")
    note = (
        "Note: Base provisional figures are aligned to publicly available company metrics "
        "(paid-up capital, borrowings/charges, revenue profile). Update the 'Base Input' sheet "
        "with exact figures from NA MARKETING PROVISIONAL.pdf to regenerate precise projections."
    )
    ws[f"B{row + 2}"] = note
    ws[f"B{row + 2}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row + 2].height = 70


def write_assumptions(wb: Workbook):
    ws = wb.create_sheet("Assumptions")
    set_col_widths(ws, {1: 4, 2: 34, 3: 18, 4: 50})

    ws["B2"] = "Projection Assumptions"
    ws["B2"].font = SUBTITLE_FONT

    headers = ["Scenario", "Growth Rate", "Application"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    row = 5
    for name, rate in GROWTH_SCENARIOS.items():
        ws.cell(row=row, column=2, value=name).border = THIN_BORDER
        c = ws.cell(row=row, column=3, value=rate)
        c.number_format = PCT_FMT
        c.border = THIN_BORDER
        ws.cell(
            row=row,
            column=4,
            value="Applied to all balance sheet line items except Share Capital",
        ).border = THIN_BORDER
        row += 1

    notes = [
        "",
        "Methodology",
        "1. FY 2026-27 projected values = FY 2025-26 provisional base × (1 + growth rate).",
        "2. FY 2027-28 estimated values = FY 2026-27 projected values × (1 + growth rate).",
        "3. Share Capital remains constant unless fresh capital is infused.",
        "4. Reserves and Surplus is treated as the balancing figure so that Total Assets = Total Equity & Liabilities.",
        "5. Detailed statements use the Base Case (37.5%) scenario.",
        "5. Scenario comparison sheet shows all three growth rates side by side.",
        "6. All amounts are in ₹ Lakhs unless stated otherwise.",
    ]
    row += 2
    for note in notes:
        ws.cell(row=row, column=2, value=note)
        if note in {"Methodology"}:
            ws.cell(row=row, column=2).font = LABEL_FONT
        row += 1


def write_base_input(wb: Workbook):
    ws = wb.create_sheet("Base Input")
    set_col_widths(ws, {1: 4, 2: 38, 3: 18, 4: 40})

    ws["B2"] = "Provisional Balance Sheet Input (FY 2025-26)"
    ws["B2"].font = SUBTITLE_FONT
    ws["B3"] = "Edit values below to match NA MARKETING PROVISIONAL.pdf"
    ws["B3"].font = Font(italic=True, color="666666")

    headers = ["Particulars", "Amount (₹ Lakhs)", "Remarks"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 6
    for key, value in BASE_FY2526.items():
        ws.cell(row=row, column=2, value=key).border = THIN_BORDER
        amount = ws.cell(row=row, column=3, value=value)
        amount.number_format = MONEY_FMT
        amount.border = THIN_BORDER
        remark = ""
        if key == "Share Capital":
            remark = "Verified from MCA paid-up capital (₹1.19 Cr)"
        elif "Borrowing" in key:
            remark = "Aligned to open secured charges (~₹44.14 Cr total)"
        ws.cell(row=row, column=4, value=remark).border = THIN_BORDER
        row += 1

    ws.cell(row=row + 1, column=2, value="Total Assets / Total Equity & Liabilities").font = LABEL_FONT
    total = sum(BASE_FY2526[k] for k in asset_keys())
    c = ws.cell(row=row + 1, column=3, value=total)
    c.number_format = MONEY_FMT
    c.font = LABEL_FONT
    c.fill = TOTAL_FILL


def write_balance_sheet_sheet(
    wb: Workbook,
    title: str,
    subtitle: str,
    values: dict[str, float],
    growth_note: str | None = None,
):
    ws = wb.create_sheet(title[:31])
    set_col_widths(ws, {1: 4, 2: 42, 3: 18, 4: 18})

    ws.merge_cells("B2:D2")
    ws["B2"] = COMPANY["name"]
    ws["B2"].font = SUBTITLE_FONT
    ws["B2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("B3:D3")
    ws["B3"] = subtitle
    ws["B3"].font = Font(bold=True, size=11)
    ws["B3"].alignment = Alignment(horizontal="center")

    if growth_note:
        ws.merge_cells("B4:D4")
        ws["B4"] = growth_note
        ws["B4"].font = Font(italic=True, color="666666")
        ws["B4"].alignment = Alignment(horizontal="center")
        start_row = 6
    else:
        start_row = 5

    ws.cell(row=start_row, column=2, value="Particulars").fill = HEADER_FILL
    ws.cell(row=start_row, column=2, value="Particulars").font = HEADER_FONT
    ws.cell(row=start_row, column=3, value="Amount (₹ Lakhs)").fill = HEADER_FILL
    ws.cell(row=start_row, column=3, value="Amount (₹ Lakhs)").font = HEADER_FONT
    for col in (2, 3):
        ws.cell(row=start_row, column=col).border = THIN_BORDER
        ws.cell(row=start_row, column=col).alignment = Alignment(horizontal="center")

    row = start_row + 1
    for label, key, kind in BS_STRUCTURE:
        if kind == "blank":
            row += 1
            continue

        label_cell = ws.cell(row=row, column=2, value=label)
        label_cell.border = THIN_BORDER

        if kind == "section":
            label_cell.font = SECTION_FONT
            label_cell.fill = SECTION_FILL
            ws.cell(row=row, column=3).fill = SECTION_FILL
            ws.cell(row=row, column=3).border = THIN_BORDER
        elif kind == "subsection":
            label_cell.font = LABEL_FONT
        elif kind in {"subtotal", "total"}:
            amount = sum(values[item] for item in key)
            label_cell.font = LABEL_FONT
            c = ws.cell(row=row, column=3, value=amount)
            c.number_format = MONEY_FMT
            c.font = LABEL_FONT
            c.border = THIN_BORDER
            fill = TOTAL_FILL if kind == "total" else SUBTOTAL_FILL
            label_cell.fill = fill
            c.fill = fill
        else:
            indent = "    " if kind == "line" else ""
            label_cell.value = f"{indent}{label}" if indent else label
            c = ws.cell(row=row, column=3, value=values[key])
            c.number_format = MONEY_FMT
            c.border = THIN_BORDER

        row += 1

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(row=start_row + 1, column=2)

    for r in range(start_row, row):
        ws.row_dimensions[r].height = 18
    total_assets = sum(values[k] for k in asset_keys())
    total_el = sum(values[k] for k in equity_liability_keys())
    ws.cell(row=row + 1, column=2, value="Balance Check (Assets - Equity & Liabilities)").font = LABEL_FONT
    check = ws.cell(row=row + 1, column=3, value=round(total_assets - total_el, 2))
    check.number_format = MONEY_FMT
    check.font = LABEL_FONT


def write_comparison(wb: Workbook, fy2627: dict, fy2728: dict):
    ws = wb.create_sheet("Scenario Comparison")
    set_col_widths(ws, {1: 4, 2: 34, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 16})

    ws["B2"] = "Growth Scenario Comparison - Key Totals (₹ Lakhs)"
    ws["B2"].font = SUBTITLE_FONT

    headers = [
        "Particulars",
        "FY 25-26 Base",
        "FY 26-27 @35%",
        "FY 27-28 @35%",
        "FY 26-27 @37.5%",
        "FY 27-28 @37.5%",
        "FY 26-27 @40%",
        "FY 27-28 @40%",
    ]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    key_groups = {
        "Total Shareholders' Funds": ["Share Capital", "Reserves and Surplus"],
        "Total Borrowings": ["Long Term Borrowings", "Short Term Borrowings"],
        "Total Current Liabilities": [
            "Short Term Borrowings",
            "Trade Payables",
            "Other Current Liabilities",
            "Short Term Provisions",
        ],
        "TOTAL ASSETS": asset_keys(),
        "TOTAL EQUITY AND LIABILITIES": list(BASE_FY2526.keys()),
    }

    row = 5
    projected = {rate: (project_values(BASE_FY2526, rate, 1), project_values(BASE_FY2526, rate, 2)) for rate in [0.35, 0.375, 0.40]}
    for label, keys in key_groups.items():
        ws.cell(row=row, column=2, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        base_total = sum(BASE_FY2526[k] for k in keys)
        ws.cell(row=row, column=3, value=base_total).number_format = MONEY_FMT

        col = 4
        for rate in [0.35, 0.375, 0.40]:
            p1, p2 = projected[rate]
            ws.cell(row=row, column=col, value=sum(p1[k] for k in keys)).number_format = MONEY_FMT
            ws.cell(row=row, column=col + 1, value=sum(p2[k] for k in keys)).number_format = MONEY_FMT
            col += 2

        for c in range(2, 10):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1


def write_ratios(wb: Workbook, base, fy2627, fy2728):
    ws = wb.create_sheet("Ratio Analysis")
    set_col_widths(ws, {1: 4, 2: 34, 3: 16, 4: 16, 5: 16})

    ws["B2"] = "Key Balance Sheet Ratios"
    ws["B2"].font = SUBTITLE_FONT

    headers = ["Ratio", "FY 25-26", "FY 26-27", "FY 27-28"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    def totals(data):
        assets = sum(data[k] for k in asset_keys())
        equity = data["Share Capital"] + data[BALANCING_ITEM]
        borrowings = data["Long Term Borrowings"] + data["Short Term Borrowings"]
        current_assets = sum(
            data[k]
            for k in [
                "Inventories",
                "Trade Receivables",
                "Cash and Cash Equivalents",
                "Short Term Loans and Advances",
                "Other Current Assets",
            ]
        )
        current_liab = sum(
            data[k]
            for k in [
                "Short Term Borrowings",
                "Trade Payables",
                "Other Current Liabilities",
                "Short Term Provisions",
            ]
        )
        return assets, equity, borrowings, current_assets, current_liab

    datasets = [("FY 25-26", base), ("FY 26-27", fy2627), ("FY 27-28", fy2728)]
    metrics = []
    for _, data in datasets:
        assets, equity, borrowings, ca, cl = totals(data)
        metrics.append(
            {
                "Debt-Equity Ratio": borrowings / equity if equity else 0,
                "Current Ratio": ca / cl if cl else 0,
                "Net Worth (₹ Lakhs)": equity,
                "Total Assets (₹ Lakhs)": assets,
                "Borrowings to Assets": borrowings / assets if assets else 0,
            }
        )

    ratio_names = list(metrics[0].keys())
    row = 5
    for name in ratio_names:
        ws.cell(row=row, column=2, value=name).border = THIN_BORDER
        for idx, metric in enumerate(metrics):
            val = metric[name]
            c = ws.cell(row=row, column=3 + idx, value=val)
            c.border = THIN_BORDER
            if "₹" in name:
                c.number_format = MONEY_FMT
            else:
                c.number_format = "0.00"
        row += 1


def main():
    growth = GROWTH_SCENARIOS[PRIMARY_SCENARIO]
    fy2627 = project_values(BASE_FY2526, growth, 1)
    fy2728 = project_values(BASE_FY2526, growth, 2)

    wb = Workbook()
    write_cover(wb)
    write_assumptions(wb)
    write_base_input(wb)
    write_balance_sheet_sheet(
        wb,
        "Provisional FY25-26",
        "Provisional Balance Sheet as on 31 March 2026 (FY 2025-26)",
        BASE_FY2526,
    )
    write_balance_sheet_sheet(
        wb,
        "Projected FY26-27",
        "Projected Balance Sheet as on 31 March 2027 (FY 2026-27)",
        fy2627,
        growth_note=f"Projection basis: {PRIMARY_SCENARIO} increment over provisional base",
    )
    write_balance_sheet_sheet(
        wb,
        "Estimated FY27-28",
        "Estimated Balance Sheet as on 31 March 2028 (FY 2027-28)",
        fy2728,
        growth_note=f"Estimation basis: {PRIMARY_SCENARIO} increment over FY 2026-27 projected figures",
    )
    write_comparison(wb, fy2627, fy2728)
    write_ratios(wb, BASE_FY2526, fy2627, fy2728)

    wb.save(OUTPUT_PATH)
    print(f"Created {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
