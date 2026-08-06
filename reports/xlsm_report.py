"""Create, load, and save Excel reports including .xlsm."""

from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltm", ".xltx"}

_XLSX_MAIN = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
_XLTX_MAIN = "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml"
_XLSM_MAIN = "application/vnd.ms-excel.sheet.macroEnabled.main+xml"
_XLTM_MAIN = "application/vnd.ms-excel.template.macroEnabled.main+xml"


def is_supported_report(path: str | Path) -> bool:
    """Return True when the path uses a supported Excel report extension."""
    return Path(path).suffix.lower() in SUPPORTED_EXTENSIONS


def create_report(
    rows: Sequence[Sequence[Any]] | None = None,
    *,
    headers: Sequence[Any] | None = None,
    sheet_name: str = "Report",
) -> WorkbookType:
    """Build an in-memory workbook suitable for saving as .xlsx or .xlsm."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    if headers:
        sheet.append(list(headers))
    if rows:
        for row in rows:
            sheet.append(list(row))

    return workbook


def load_report(path: str | Path, *, data_only: bool = False) -> WorkbookType:
    """
    Load an Excel report.

    For .xlsm/.xltm files, VBA macros are preserved when the workbook is saved
    again (openpyxl keep_vba=True). Macros are not executed by this library.
    """
    path = Path(path)
    if not is_supported_report(path):
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Unsupported report type: {path.suffix!r}. Supported: {supported}")

    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    return load_workbook(path, keep_vba=keep_vba, data_only=data_only)


def _ensure_macro_enabled_content_types(path: Path) -> None:
    """
    openpyxl only emits macro-enabled content types when a VBA archive exists.
    For new .xlsm/.xltm reports without VBA, rewrite the workbook content type
    so Excel treats the file as a macro-enabled package.
    """
    suffix = path.suffix.lower()
    if suffix == ".xlsm":
        replacements = {
            _XLSX_MAIN: _XLSM_MAIN,
            _XLTX_MAIN: _XLSM_MAIN,
        }
    elif suffix == ".xltm":
        replacements = {
            _XLSX_MAIN: _XLTM_MAIN,
            _XLTX_MAIN: _XLTM_MAIN,
        }
    else:
        return

    buffer = io.BytesIO()
    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            data = src.read(info.filename)
            if info.filename == "[Content_Types].xml":
                text = data.decode("utf-8")
                for old, new in replacements.items():
                    text = text.replace(old, new)
                data = text.encode("utf-8")
            dst.writestr(info, data)

    path.write_bytes(buffer.getvalue())


def save_report(workbook: WorkbookType, path: str | Path) -> Path:
    """
    Save a workbook to disk.

    Paths ending in .xlsm are written as macro-enabled workbooks. If the
    workbook was loaded from an existing .xlsm with keep_vba=True, macros are
    retained. Newly created workbooks saved as .xlsm use the macro-enabled
    package content type (no VBA project unless one was loaded).
    """
    path = Path(path)
    if not is_supported_report(path):
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Unsupported report type: {path.suffix!r}. Supported: {supported}")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)

    if path.suffix.lower() in {".xlsm", ".xltm"} and not getattr(workbook, "vba_archive", None):
        _ensure_macro_enabled_content_types(path)

    return path


def read_sheet_rows(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    data_only: bool = True,
) -> list[tuple[Any, ...]]:
    """Read all rows from a sheet in a supported Excel report."""
    workbook = load_report(path, data_only=data_only)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    return [tuple(cell.value for cell in row) for row in sheet.iter_rows()]


def write_report_from_dicts(
    path: str | Path,
    records: Iterable[Mapping[str, Any]],
    *,
    sheet_name: str = "Report",
    fieldnames: Sequence[str] | None = None,
) -> Path:
    """Write mapping records to an Excel report (.xlsx or .xlsm)."""
    records = list(records)
    if fieldnames is None:
        fieldnames = list(records[0].keys()) if records else []

    rows = [[record.get(name) for name in fieldnames] for record in records]
    workbook = create_report(rows, headers=fieldnames, sheet_name=sheet_name)
    return save_report(workbook, path)
