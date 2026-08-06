#!/usr/bin/env python3
"""
RetailX Uploader Builder
------------------------
Dynamically builds the Uploader sheet from Extract + GL Master.

Fixes the VLOOKUP/SUMIFS limitation: new states and product categories
appearing in Extract are discovered automatically and posted — no hardcoded
state×category grid required.

Usage:
  python3 build_retailx_uploader.py
  python3 build_retailx_uploader.py --input RetailX.xlsx --output RetailX.xlsx
"""

from __future__ import annotations

import argparse
import calendar
from collections import defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants / config defaults
# ---------------------------------------------------------------------------

STATE_ALIAS = {"IN-GJ": "IN-GR", "IN-OR": "IN-OS"}

DEBTOR_POSTPAID = 131144
DEBTOR_PREPAID = 131102
DEBTOR_PBO = 131126

IGST_GL = {5: 225001, 12: 225002, 18: 225003, 28: 225004}
CGST_GL = {5: 225006, 12: 225007, 18: 225008, 28: 225009}
SGST_GL = {5: 225011, 12: 225012, 18: 225013, 28: 225014}

# Voucher series (configurable via Config sheet when present)
DEFAULT_VOUCHERS = [
    {
        "voucher_no": 11,
        "key": "SALES",
        "report_names": ["NONDIGITAL"],
        "narration": "Being sales booked for the month of {month}",
        "function": "Sales",
    },
    {
        "voucher_no": 12,
        "key": "SALES_RETURN",
        "report_names": ["RETURN_CREATED"],
        "narration": "Being Sales Return booked for the month of {month}",
        "function": "Sales",
    },
    {
        "voucher_no": 13,
        "key": "SHIPPING",
        "report_names": ["NONDIGITAL", "RETURN_CREATED"],
        "narration": "Being Shipping Revenue booked for the month of {month}",
        "function": "Sales",
    },
    {
        "voucher_no": 14,
        "key": "PBO",
        "report_names": ["PBO_SALES", "PBO_RETURN"],
        "narration": "Being PBO Revenue booked for the month of {month}",
        "function": "Sales",
    },
    {
        "voucher_no": 15,
        "key": "PRICE_DROP",
        "report_names": ["PRICE_DROP"],
        "narration": "Price drop for the month of {month}",
        "function": "Sales",
    },
    {
        "voucher_no": 16,
        "key": "BUYER_FEE",
        "report_names": ["BUYER_FEE"],
        "narration": "Being Secure Packaging revenue booked for the month {month}",
        "function": "Sales",
    },
    {
        "voucher_no": 17,
        "key": "PREXO_BUMPUP",
        "report_names": ["PREXO_BUMPUP"],
        "narration": "Being PREXO BUMPUP Revenue booked for the month of {month}",
        "function": "Sales",
    },
]

UPLOADER_HEADERS = [
    "VoucherType",
    "Account Name",
    "Date",
    "Ref New Field",
    "Ledger Narration",
    "Voucher No",
    " State Name",
    "Function",
    "Location",
    "Debit Amount",
    "Credit Amount",
    "Narration",
    "Sl no",
    "Invoice  Type",
    "Company Code",
    "Is provision reverse",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def r2(v: float) -> float:
    return round(float(v or 0.0), 2)


def alias_state(state: str | None) -> str:
    if not state:
        return "IN-OTH"
    return STATE_ALIAS.get(str(state).strip(), str(state).strip())


def snap_gst_rate(tax: float, base: float, available=(5, 18)) -> int:
    """Infer GST slab when Extract has no rate columns."""
    if abs(base) < 1e-9:
        return 18
    pct = abs(tax / base) * 100.0
    opts = tuple(available)
    # Empirically best split for RetailX extracts that only use 5% / 18%
    if set(opts) <= {5, 18} or opts == (5, 18):
        return 5 if pct < 8.0 else 18
    return min(opts, key=lambda x: abs(x - pct))


def parse_month_end(month_label: str) -> datetime:
    """Parse labels like Jun'26 into month-end date."""
    text = str(month_label).strip().replace("'", "").replace("’", "")
    # Try MonYY / MonYYYY
    for fmt in ("%b%y", "%b%Y", "%B%y", "%B%Y"):
        try:
            dt = datetime.strptime(text, fmt)
            last = calendar.monthrange(dt.year, dt.month)[1]
            return datetime(dt.year, dt.month, last)
        except ValueError:
            continue
    # Fallback: today month-end previous
    today = datetime.today()
    y, m = (today.year, today.month - 1) if today.month > 1 else (today.year - 1, 12)
    return datetime(y, m, calendar.monthrange(y, m)[1])


def is_tax_or_debtor_key(prod: str) -> bool:
    p = str(prod or "").strip().upper()
    if not p:
        return True
    if p in {"POSTPAID", "PREPAID", "BNPL", "DEBTOR"}:
        return True
    return any(p.startswith(x) for x in ("CGST", "SGST", "IGST", "CGST", "SGST"))


# ---------------------------------------------------------------------------
# GL Master lookups
# ---------------------------------------------------------------------------

def build_gl_maps(ws_gl):
    """Return lookup dicts from GL Master."""
    rev = {}          # category -> GL (Sales, Rec blank)
    disc = {}         # category -> discount GL
    ret = {}          # category -> Sales Return GL
    ret_disc = {}     # category -> return discount GL
    special = defaultdict(dict)  # classification -> product -> GL

    for r in range(2, ws_gl.max_row + 1):
        clas = ws_gl.cell(r, 1).value
        prod = ws_gl.cell(r, 2).value
        rec = ws_gl.cell(r, 3).value
        gl = ws_gl.cell(r, 5).value
        if clas is None or gl is None:
            continue
        clas_s = str(clas).strip()
        prod_s = None if prod is None else str(prod).strip()
        rec_s = None if rec is None else str(rec).strip()

        special[clas_s][prod_s if prod_s is not None else ""] = gl

        if clas_s == "Sales" and prod_s and not is_tax_or_debtor_key(prod_s):
            if rec_s and "discount" in rec_s.lower():
                disc[prod_s] = gl
            elif rec_s is None:
                rev[prod_s] = gl

        if clas_s == "Sales Return" and prod_s and not is_tax_or_debtor_key(prod_s):
            if rec_s and "discount" in rec_s.lower():
                ret_disc[prod_s] = gl
            elif rec_s is None:
                ret[prod_s] = gl

    # Discount fallback: Sales Return Discounts often hold the same keys
    for k, v in ret_disc.items():
        disc.setdefault(k, v)

    return {
        "rev": rev,
        "disc": disc,
        "ret": ret,
        "ret_disc": ret_disc,
        "special": special,
    }


def gl_special(maps, classification: str, product: str | None, default=None):
    bucket = maps["special"].get(classification, {})
    if product is None:
        return bucket.get("", default)
    # try exact then case-insensitive
    if product in bucket:
        return bucket[product]
    for k, v in bucket.items():
        if k is not None and str(k).lower() == str(product).lower():
            return v
    return default


# ---------------------------------------------------------------------------
# Extract reader
# ---------------------------------------------------------------------------

def read_extract(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers, 1) if h}
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, i).value for h, i in idx.items()}
        if row.get("report_name") is None:
            continue
        rows.append(row)
    return headers, idx, rows


def amt(row, col) -> float:
    return float(row.get(col) or 0.0)


def has_rate_cols(headers) -> bool:
    names = {str(h).lower() for h in headers if h}
    return {"cgst_rate", "igst_rate"} & names or {"cgst_rate", "sgst_utgst_rate", "igst_rate"} <= names


def row_gst_rate(row, base_col_value: float) -> int:
    """Use explicit rates when present; else infer from tax/base."""
    # Prefer IGST rate if > 0 else CGST*2
    for key in ("igst_rate", "IGST_RATE"):
        if row.get(key) not in (None, "", 0, 0.0, "0", "0.0"):
            try:
                v = float(row[key])
                if v > 0:
                    return snap_gst_rate(v, 100, (5, 12, 18, 28)) if v <= 40 else int(v)
            except (TypeError, ValueError):
                pass
    for key in ("cgst_rate", "CGST_RATE"):
        if row.get(key) not in (None, "", 0, 0.0, "0", "0.0"):
            try:
                v = float(row[key])
                if v > 0:
                    return snap_gst_rate(v * 2, 100, (5, 12, 18, 28))
            except (TypeError, ValueError):
                pass
    # Infer — tax columns vary by report; caller passes meaningful base
    return snap_gst_rate(0, base_col_value)  # placeholder; callers pass tax too


def classify_tax(tax: float, base: float, state_from, state_to, rate: int | None = None):
    """Yield (gl, state_alias, amount) tax postings."""
    if abs(tax) < 1e-9:
        return []
    rate = rate if rate is not None else snap_gst_rate(tax, base)
    # Clamp to known GLs
    if rate not in IGST_GL:
        rate = snap_gst_rate(rate, 100, tuple(IGST_GL.keys()))
    st = alias_state(state_from)
    intra = str(state_from) == str(state_to)
    out = []
    if intra:
        out.append((CGST_GL[rate], st, tax / 2.0))
        out.append((SGST_GL[rate], st, tax / 2.0))
    else:
        out.append((IGST_GL[rate], st, tax))
    return out


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

class UploaderBuilder:
    def __init__(self, company_code="UTSPLFK", invoice_type="B2C", location="Business Operations"):
        self.company_code = company_code
        self.invoice_type = invoice_type
        self.location = location
        self.rows = []
        self.unmapped = []
        self.sl = 0

    def add(
        self,
        *,
        voucher_no,
        account,
        date,
        state,
        function,
        debit,
        credit,
        narration,
        voucher_type="AR-JV Sale",
    ):
        debit, credit = r2(debit), r2(credit)
        if abs(debit) < 0.005 and abs(credit) < 0.005:
            return
        # Keep signs clean: only one side
        if debit < 0 and credit == 0:
            credit, debit = -debit, 0.0
        if credit < 0 and debit == 0:
            debit, credit = -credit, 0.0
        self.sl += 1
        self.rows.append(
            {
                "VoucherType": voucher_type,
                "Account Name": account,
                "Date": date,
                "Ref New Field": None,
                "Ledger Narration": None,
                "Voucher No": voucher_no,
                " State Name": state,
                "Function": function,
                "Location": self.location,
                "Debit Amount": r2(debit),
                "Credit Amount": r2(credit),
                "Narration": narration,
                "Sl no": self.sl,
                "Invoice  Type": self.invoice_type,
                "Company Code": self.company_code,
                "Is provision reverse": "No",
            }
        )

    def note_unmapped(self, kind, key, detail=""):
        self.unmapped.append({"Type": kind, "Key": key, "Detail": detail})


# ---------------------------------------------------------------------------
# Voucher generators
# ---------------------------------------------------------------------------

def gen_sales(b: UploaderBuilder, rows, maps, voucher_no, date, narration, function):
    debtor = defaultdict(float)
    rev_agg = defaultdict(float)      # (gl, state) -> credit
    disc_agg = defaultdict(float)     # (gl, state) -> debit
    tax_agg = defaultdict(float)      # (gl, state) -> credit

    for row in rows:
        if row.get("report_name") != "NONDIGITAL":
            continue
        cat = str(row.get("analytics_category") or "").strip()
        st_from = row.get("state_code_from")
        st_to = row.get("state_code_to")
        sc = str(row.get("sc_type") or "").lower()
        rev = amt(row, "Sales Revenue")
        disc = amt(row, "Sales Discount")
        tax = amt(row, "Sales Tax")

        debtor[sc] += rev + disc + tax

        rg = maps["rev"].get(cat)
        dg = maps["disc"].get(cat)
        if rg is None and abs(rev) > 0.005:
            b.note_unmapped("Sales Revenue GL", cat, "Classification=Sales, Rec blank")
        if dg is None and abs(disc) > 0.005:
            b.note_unmapped("Sales Discount GL", cat, "Classification=Sales/Sales Return Discounts")

        st = alias_state(st_from)
        if rg:
            rev_agg[(rg, st)] += rev
        if dg:
            disc_agg[(dg, st)] += abs(disc)
        for gl, st_a, tax_amt in classify_tax(tax, rev, st_from, st_to):
            tax_agg[(gl, st_a)] += tax_amt

    # Debtors
    b.add(voucher_no=voucher_no, account=DEBTOR_POSTPAID, date=date, state="IN-OTH",
          function=function, debit=debtor.get("postpaid", 0), credit=0, narration=narration)
    b.add(voucher_no=voucher_no, account=DEBTOR_PREPAID, date=date, state="IN-OTH",
          function=function, debit=debtor.get("prepaid", 0), credit=0, narration=narration)

    for (gl, st), v in sorted(rev_agg.items(), key=lambda x: (str(x[0][0]), x[0][1])):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)
    for (gl, st), v in sorted(disc_agg.items(), key=lambda x: (str(x[0][0]), x[0][1])):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=v, credit=0, narration=narration)
    for (gl, st), v in sorted(tax_agg.items(), key=lambda x: (str(x[0][0]), x[0][1])):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)


def gen_sales_return(b: UploaderBuilder, rows, maps, voucher_no, date, narration, function):
    debtor = defaultdict(float)
    ret_agg = defaultdict(float)       # debit abs(sales return)
    disc_agg = defaultdict(float)      # credit return discount
    tax_agg = defaultdict(float)       # debit abs(return tax)

    for row in rows:
        if row.get("report_name") != "RETURN_CREATED":
            continue
        cat = str(row.get("analytics_category") or "").strip()
        st_from = row.get("state_code_from")
        st_to = row.get("state_code_to")
        sc = str(row.get("sc_type") or "").lower()
        sret = amt(row, "Sales Return")
        rdisc = amt(row, "Return Discount")
        rtax = amt(row, "Return Tax")

        debtor[sc] += sret + rdisc + rtax  # typically negative net → credit

        rg = maps["ret"].get(cat)
        dg = maps["disc"].get(cat) or maps["ret_disc"].get(cat)
        if rg is None and abs(sret) > 0.005:
            b.note_unmapped("Sales Return GL", cat, "Classification=Sales Return, Rec blank")
        st = alias_state(st_from)
        if rg:
            ret_agg[(rg, st)] += abs(sret)
        if dg:
            disc_agg[(dg, st)] += abs(rdisc) if rdisc > 0 else abs(rdisc)
        for gl, st_a, tax_amt in classify_tax(rtax, sret, st_from, st_to):
            tax_agg[(gl, st_a)] += tax_amt

    # Debtors — credit when net negative
    for sc, gl in (("postpaid", DEBTOR_POSTPAID), ("prepaid", DEBTOR_PREPAID)):
        net = debtor.get(sc, 0)
        # net usually negative → credit abs
        b.add(voucher_no=voucher_no, account=gl, date=date, state="IN-OTH",
              function=function, debit=net if net > 0 else 0,
              credit=(-net if net < 0 else 0), narration=narration)

    for (gl, st), v in sorted(ret_agg.items(), key=lambda x: (str(x[0][0]), x[0][1])):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=v, credit=0, narration=narration)
    for (gl, st), v in sorted(disc_agg.items(), key=lambda x: (str(x[0][0]), x[0][1])):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)
    for (gl, st), v in sorted(tax_agg.items(), key=lambda x: (str(x[0][0]), x[0][1])):
        # Return tax is negative in Extract → post absolute as Debit
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=abs(v), credit=0, narration=narration)


def gen_shipping(b: UploaderBuilder, rows, maps, voucher_no, date, narration, function):
    ship_gl = gl_special(maps, "Shipping Revenue", "regular", 401051)
    debtor = defaultdict(float)
    # Separate sales vs return legs so both directions stay visible
    rev_cr = defaultdict(float)
    rev_dr = defaultdict(float)
    tax_cr = defaultdict(float)
    tax_dr = defaultdict(float)

    for row in rows:
        report = row.get("report_name")
        if report not in ("NONDIGITAL", "RETURN_CREATED"):
            continue
        ship = amt(row, "Shipping Amount")
        stax = amt(row, "Shipping Tax")
        if abs(ship) < 1e-9 and abs(stax) < 1e-9:
            continue
        sc = str(row.get("sc_type") or "").lower()
        st_from = row.get("state_code_from")
        st_to = row.get("state_code_to")
        st = alias_state(st_from)

        debtor[sc] += ship + stax

        if ship >= 0:
            rev_cr[(ship_gl, st)] += ship
        else:
            rev_dr[(ship_gl, st)] += abs(ship)

        for gl, st_a, tax_amt in classify_tax(stax, ship, st_from, st_to):
            if tax_amt >= 0:
                tax_cr[(gl, st_a)] += tax_amt
            else:
                tax_dr[(gl, st_a)] += abs(tax_amt)

    b.add(voucher_no=voucher_no, account=DEBTOR_POSTPAID, date=date, state="IN-OTH",
          function=function, debit=debtor.get("postpaid", 0), credit=0, narration=narration)
    b.add(voucher_no=voucher_no, account=DEBTOR_PREPAID, date=date, state="IN-OTH",
          function=function, debit=debtor.get("prepaid", 0), credit=0, narration=narration)

    for (gl, st), v in sorted(rev_cr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)
    for (gl, st), v in sorted(tax_cr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)
    for (gl, st), v in sorted(rev_dr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=v, credit=0, narration=narration)
    for (gl, st), v in sorted(tax_dr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=v, credit=0, narration=narration)


def gen_pbo(b: UploaderBuilder, rows, maps, voucher_no, date, narration, function):
    rev_gl = gl_special(maps, "PBO Sales", "Mobile", 400169)
    debtor_gl = gl_special(maps, "PBO Sales", "debtor", DEBTOR_PBO)
    debtor = 0.0
    rev_cr = defaultdict(float)
    rev_dr = defaultdict(float)
    tax_cr = defaultdict(float)
    tax_dr = defaultdict(float)

    for row in rows:
        report = row.get("report_name")
        if report not in ("PBO_SALES", "PBO_RETURN"):
            continue
        rev = amt(row, "PBO Revenue")
        tax = amt(row, "PBO TAX")
        if abs(rev) < 1e-9 and abs(tax) < 1e-9:
            continue
        st_from = row.get("state_code_from")
        st_to = row.get("state_code_to")
        st = alias_state(st_from)
        debtor += rev + tax

        if rev >= 0:
            rev_cr[(rev_gl, st)] += rev
        else:
            rev_dr[(rev_gl, st)] += abs(rev)

        for gl, st_a, tax_amt in classify_tax(tax, rev, st_from, st_to):
            if tax_amt >= 0:
                tax_cr[(gl, st_a)] += tax_amt
            else:
                tax_dr[(gl, st_a)] += abs(tax_amt)

    b.add(voucher_no=voucher_no, account=debtor_gl, date=date, state="IN-OTH",
          function=function, debit=debtor if debtor > 0 else 0,
          credit=(-debtor if debtor < 0 else 0), narration=narration)

    for (gl, st), v in sorted(rev_cr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)
    for (gl, st), v in sorted(tax_cr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)
    for (gl, st), v in sorted(rev_dr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=v, credit=0, narration=narration)
    for (gl, st), v in sorted(tax_dr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=v, credit=0, narration=narration)


def gen_price_drop(b: UploaderBuilder, rows, maps, voucher_no, date, narration, function):
    debtor = defaultdict(float)
    rev_agg = defaultdict(float)
    tax_agg = defaultdict(float)

    for row in rows:
        if row.get("report_name") != "PRICE_DROP":
            continue
        cat = str(row.get("analytics_category") or "").strip()
        # numeric / unknown cats → Mobile revenue GL (matches existing file)
        rg = maps["rev"].get(cat) or maps["rev"].get("Mobile") or 401121
        if cat not in maps["rev"] and not cat.isdigit():
            b.note_unmapped("Price Drop Revenue GL", cat, "fell back to Mobile")
        sc = str(row.get("sc_type") or "").lower()
        st_from = row.get("state_code_from")
        st_to = row.get("state_code_to")
        st = alias_state(st_from)
        pd = amt(row, "PRICE DROP")
        pdt = amt(row, "PRICE DROP TAX")
        debtor[sc] += pd + pdt
        rev_agg[(rg, st)] += pd
        for gl, st_a, tax_amt in classify_tax(pdt, pd, st_from, st_to):
            tax_agg[(gl, st_a)] += tax_amt

    for sc, gl in (("postpaid", DEBTOR_POSTPAID), ("prepaid", DEBTOR_PREPAID)):
        net = debtor.get(sc, 0)
        if abs(net) < 0.005:
            continue
        b.add(voucher_no=voucher_no, account=gl, date=date, state="IN-OTH",
              function=function, debit=net if net > 0 else 0,
              credit=(-net if net < 0 else 0), narration=narration)

    for (gl, st), v in sorted(rev_agg.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=abs(v), credit=0, narration=narration)
    for (gl, st), v in sorted(tax_agg.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=abs(v), credit=0, narration=narration)


def gen_buyer_fee(b: UploaderBuilder, rows, maps, voucher_no, date, narration, function):
    rev_gl = gl_special(maps, "Secure Packaging Revenue", None, 401056)
    if rev_gl is None:
        rev_gl = gl_special(maps, "Prexo Pickup Revenue", None, 401056) or 401056
    debtor = defaultdict(float)
    rev_cr = defaultdict(float)
    rev_dr = defaultdict(float)
    tax_cr = defaultdict(float)
    tax_dr = defaultdict(float)

    for row in rows:
        if row.get("report_name") != "BUYER_FEE":
            continue
        sc = str(row.get("sc_type") or "").lower()
        st_from = row.get("state_code_from")
        st_to = row.get("state_code_to")
        st = alias_state(st_from)
        fee = amt(row, "Buyer Fees")
        tax = amt(row, "BuyerTax")
        debtor[sc] += fee + tax
        if fee >= 0:
            rev_cr[(rev_gl, st)] += fee
        else:
            rev_dr[(rev_gl, st)] += abs(fee)
        for gl, st_a, tax_amt in classify_tax(tax, fee, st_from, st_to):
            if tax_amt >= 0:
                tax_cr[(gl, st_a)] += tax_amt
            else:
                tax_dr[(gl, st_a)] += abs(tax_amt)

    b.add(voucher_no=voucher_no, account=DEBTOR_POSTPAID, date=date, state="IN-OTH",
          function=function, debit=debtor.get("postpaid", 0), credit=0, narration=narration)
    b.add(voucher_no=voucher_no, account=DEBTOR_PREPAID, date=date, state="IN-OTH",
          function=function, debit=debtor.get("prepaid", 0), credit=0, narration=narration)

    for (gl, st), v in sorted(rev_cr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)
    for (gl, st), v in sorted(tax_cr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)
    for (gl, st), v in sorted(rev_dr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=v, credit=0, narration=narration)
    for (gl, st), v in sorted(tax_dr.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=v, credit=0, narration=narration)


def gen_prexo_bumpup(b: UploaderBuilder, rows, maps, voucher_no, date, narration, function):
    rev_gl = gl_special(maps, "PREXO BUMPUP", "Mobile", 400169)
    debtor_gl = gl_special(maps, "PREXO BUMPUP", "debtor", DEBTOR_PBO)
    debtor = 0.0
    rev_agg = defaultdict(float)
    tax_agg = defaultdict(float)

    for row in rows:
        if row.get("report_name") != "PREXO_BUMPUP":
            continue
        st_from = row.get("state_code_from")
        st_to = row.get("state_code_to")
        st = alias_state(st_from)
        rev = amt(row, "PREXOBUMPUP")
        tax = amt(row, "PREXOBUMP_TAX")
        debtor += rev + tax
        rev_agg[(rev_gl, st)] += rev
        for gl, st_a, tax_amt in classify_tax(tax, rev, st_from, st_to):
            tax_agg[(gl, st_a)] += tax_amt

    b.add(voucher_no=voucher_no, account=debtor_gl, date=date, state="IN-OTH",
          function=function, debit=debtor if debtor > 0 else 0,
          credit=(-debtor if debtor < 0 else 0), narration=narration)
    for (gl, st), v in sorted(rev_agg.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)
    for (gl, st), v in sorted(tax_agg.items()):
        b.add(voucher_no=voucher_no, account=gl, date=date, state=st, function=function,
              debit=0, credit=v, narration=narration)


GENERATORS = {
    "SALES": gen_sales,
    "SALES_RETURN": gen_sales_return,
    "SHIPPING": gen_shipping,
    "PBO": gen_pbo,
    "PRICE_DROP": gen_price_drop,
    "BUYER_FEE": gen_buyer_fee,
    "PREXO_BUMPUP": gen_prexo_bumpup,
}


# ---------------------------------------------------------------------------
# Workbook writers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def write_sheet_rows(ws, headers, rows, clear=True):
    if clear:
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        # ensure headers
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            val = row.get(h)
            cell = ws.cell(i, c, val)
            cell.border = THIN
            if h == "Date" and isinstance(val, datetime):
                cell.number_format = "DD-MMM-YYYY"
            if h in ("Debit Amount", "Credit Amount") and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"


def ensure_sheet(wb, name, headers=None):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if headers:
        for c, h in enumerate(headers, 1):
            if ws.cell(1, c).value != h:
                ws.cell(1, c, h)
    return ws


def write_config(ws, vouchers, month_label, date):
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = "RetailX Uploader Config"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = "Month Label"
    ws["B2"] = month_label
    ws["A3"] = "Booking Date"
    ws["B3"] = date
    ws["B3"].number_format = "DD-MMM-YYYY"
    ws["A4"] = "Company Code"
    ws["B4"] = "UTSPLFK"
    ws["A5"] = "Invoice Type"
    ws["B5"] = "B2C"
    ws["A6"] = "State Alias (Extract → Uploader)"
    ws["B6"] = "IN-GJ→IN-GR ; IN-OR→IN-OS"

    ws["A8"] = "Voucher No"
    ws["B8"] = "Key"
    ws["C8"] = "Report Name(s)"
    ws["D8"] = "Narration template"
    ws["E8"] = "Function"
    for c in range(1, 6):
        ws.cell(8, c).fill = HEADER_FILL
        ws.cell(8, c).font = HEADER_FONT

    for i, v in enumerate(vouchers, 9):
        ws.cell(i, 1, v["voucher_no"])
        ws.cell(i, 2, v["key"])
        ws.cell(i, 3, ", ".join(v["report_names"]))
        ws.cell(i, 4, v["narration"])
        ws.cell(i, 5, v["function"])

    ws["A18"] = "How regeneration works"
    ws["A18"].font = Font(bold=True)
    ws["A19"] = (
        "1) Paste latest Extract data into the Extract sheet (keep header row). "
        "2) Ensure GL Master has GLs for every analytics_category under Sales / Sales Return. "
        "3) Run: python3 build_retailx_uploader.py   OR import VBA module GenerateUploader and run macro. "
        "New states and categories are picked up automatically — no VLOOKUP grid to maintain."
    )
    ws["A19"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A19:E22")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 14


def write_instructions(ws):
    ws.delete_rows(1, ws.max_row)
    lines = [
        ("RetailX Dynamic Uploader — Instructions", True),
        ("", False),
        ("Problem solved", True),
        ("Earlier Uploader used hardcoded SUMIFS/VLOOKUP rows per state × category. When Extract had a new state or product category, those amounts were missed.", False),
        ("This rebuild discovers every state and category present in Extract and posts them using GL Master lookups.", False),
        ("", False),
        ("Voucher design (separate series per revenue type)", True),
        ("11  Sales              ← report_name = NONDIGITAL (Revenue + Discount + Tax + Debtors)", False),
        ("12  Sales Return       ← report_name = RETURN_CREATED", False),
        ("13  Shipping Revenue   ← Shipping Amount/Tax on NONDIGITAL + RETURN_CREATED", False),
        ("14  PBO Revenue        ← PBO_SALES + PBO_RETURN (net debtor)", False),
        ("15  Price Drop         ← PRICE_DROP", False),
        ("16  Secure Packaging   ← BUYER_FEE (Buyer Fees + BuyerTax)", False),
        ("17  PREXO BUMPUP       ← PREXO_BUMPUP", False),
        ("", False),
        ("GL lookup rules", True),
        ("Sales revenue GL      = GL Master where Classification='Sales' and Rec blank, Product=analytics_category", False),
        ("Sales discount GL     = Classification='Sales' or 'Sales Return' with Rec containing 'Discounts'", False),
        ("Sales return GL       = Classification='Sales Return', Rec blank", False),
        ("Tax GLs               = IGST/CGST/SGST by rate; intra-state (from=to) splits CGST+SGST, else IGST", False),
        ("State aliases         = IN-GJ posted as IN-GR; IN-OR posted as IN-OS (Uploader convention)", False),
        ("", False),
        ("How to refresh after a new Extract", True),
        ("Option A — Python (recommended on any OS):  python3 build_retailx_uploader.py", False),
        ("Option B — Excel VBA (Windows/Mac Excel): Alt+F11 → File → Import File → GenerateUploader.bas → Run GenerateUploader", False),
        ("", False),
        ("Check Unmapped sheet after each run. Add any missing category GLs into GL Master, then regenerate.", False),
        ("Balance check: for each Voucher No, sum(Debit) should equal sum(Credit) (minor rounding < ₹1 may remain).", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        ws.cell(i, 1, text)
        if bold:
            ws.cell(i, 1).font = Font(bold=True, color="1F4E79", size=12 if i == 1 else 11)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 120


def write_control_totals(ws, builder_rows):
    ws.delete_rows(1, ws.max_row)
    headers = ["Voucher No", "Narration", "Debit", "Credit", "Difference"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    by_v = defaultdict(lambda: {"d": 0.0, "c": 0.0, "n": ""})
    for row in builder_rows:
        v = row["Voucher No"]
        by_v[v]["d"] += row["Debit Amount"] or 0
        by_v[v]["c"] += row["Credit Amount"] or 0
        by_v[v]["n"] = row["Narration"]

    for i, v in enumerate(sorted(by_v), 2):
        ws.cell(i, 1, v)
        ws.cell(i, 2, by_v[v]["n"])
        ws.cell(i, 3, r2(by_v[v]["d"])).number_format = "#,##0.00"
        ws.cell(i, 4, r2(by_v[v]["c"])).number_format = "#,##0.00"
        diff = r2(by_v[v]["d"] - by_v[v]["c"])
        cell = ws.cell(i, 5, diff)
        cell.number_format = "#,##0.00"
        if abs(diff) > 1:
            cell.font = Font(color="9C0006", bold=True)

    for col, w in enumerate([14, 60, 18, 18, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def balance_voucher_rows(rows, tol=0.05):
    """Push residual rounding difference onto the largest debtor line of each voucher."""
    by_v = defaultdict(list)
    for i, row in enumerate(rows):
        by_v[row["Voucher No"]].append(i)
    for v, idxs in by_v.items():
        deb = sum(rows[i]["Debit Amount"] or 0 for i in idxs)
        cred = sum(rows[i]["Credit Amount"] or 0 for i in idxs)
        diff = r2(deb - cred)
        if abs(diff) < 0.005 or abs(diff) > 50:
            continue
        # Prefer IN-OTH debtor lines
        candidates = [
            i for i in idxs
            if rows[i][" State Name"] == "IN-OTH" and (rows[i]["Debit Amount"] or rows[i]["Credit Amount"])
        ] or idxs
        # Adjust the largest absolute amount line on the heavy side
        if diff > 0:
            # debit heavy → reduce largest debit / increase largest credit
            i = max(candidates, key=lambda x: rows[x]["Credit Amount"] or 0)
            if (rows[i]["Credit Amount"] or 0) > 0:
                rows[i]["Credit Amount"] = r2(rows[i]["Credit Amount"] + diff)
            else:
                i = max(candidates, key=lambda x: rows[x]["Debit Amount"] or 0)
                rows[i]["Debit Amount"] = r2(rows[i]["Debit Amount"] - diff)
        else:
            i = max(candidates, key=lambda x: rows[x]["Debit Amount"] or 0)
            if (rows[i]["Debit Amount"] or 0) > 0:
                rows[i]["Debit Amount"] = r2(rows[i]["Debit Amount"] - diff)
            else:
                i = max(candidates, key=lambda x: rows[x]["Credit Amount"] or 0)
                rows[i]["Credit Amount"] = r2(rows[i]["Credit Amount"] + (-diff))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build(input_path: Path, output_path: Path):
    wb = load_workbook(input_path)
    if "Extract" not in wb.sheetnames or "GL Master" not in wb.sheetnames:
        raise SystemExit("Workbook must contain Extract and GL Master sheets")

    headers, idx, extract_rows = read_extract(wb["Extract"])
    if not extract_rows:
        raise SystemExit("Extract sheet has no data rows")

    month_label = extract_rows[0].get("Month") or "Unknown"
    book_date = parse_month_end(month_label)
    maps = build_gl_maps(wb["GL Master"])

    present_reports = sorted({r.get("report_name") for r in extract_rows if r.get("report_name")})
    print(f"Month: {month_label} | Date: {book_date.date()} | Reports: {present_reports}")
    print(f"Extract rows: {len(extract_rows)} | Sales GLs: {len(maps['rev'])} | Return GLs: {len(maps['ret'])}")

    builder = UploaderBuilder()
    vouchers = DEFAULT_VOUCHERS

    for spec in vouchers:
        # Skip voucher if none of its reports exist in extract
        if not any(r in present_reports for r in spec["report_names"]):
            print(f"  skip voucher {spec['voucher_no']} ({spec['key']}) — reports not in Extract")
            continue
        narr = spec["narration"].format(month=month_label)
        gen = GENERATORS[spec["key"]]
        before = len(builder.rows)
        gen(builder, extract_rows, maps, spec["voucher_no"], book_date, narr, spec["function"])
        print(f"  voucher {spec['voucher_no']} {spec['key']}: {len(builder.rows) - before} lines")

    balance_voucher_rows(builder.rows)

    # Uploader sheet
    ws_up = ensure_sheet(wb, "Uploader")
    if ws_up.max_row >= 1:
        ws_up.delete_rows(1, ws_up.max_row)
    for c, h in enumerate(UPLOADER_HEADERS, 1):
        cell = ws_up.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, row in enumerate(builder.rows, 2):
        for c, h in enumerate(UPLOADER_HEADERS, 1):
            val = row.get(h)
            cell = ws_up.cell(i, c, val)
            cell.border = THIN
            if h == "Date" and isinstance(val, datetime):
                cell.number_format = "DD-MMM-YYYY"
            if h in ("Debit Amount", "Credit Amount"):
                cell.number_format = "#,##0.00"

    widths = [12, 14, 12, 12, 14, 12, 12, 10, 18, 14, 14, 50, 8, 12, 12, 16]
    for i, w in enumerate(widths, 1):
        ws_up.column_dimensions[get_column_letter(i)].width = w

    write_config(ensure_sheet(wb, "Config"), vouchers, month_label, book_date)
    write_instructions(ensure_sheet(wb, "Instructions"))
    write_control_totals(ensure_sheet(wb, "Control"), builder.rows)

    ws_un = ensure_sheet(wb, "Unmapped")
    seen = set()
    uniq = []
    for u in builder.unmapped:
        key = (u["Type"], u["Key"])
        if key in seen:
            continue
        seen.add(key)
        uniq.append(u)
    if ws_un.max_row >= 1:
        ws_un.delete_rows(1, ws_un.max_row)
    write_sheet_rows(ws_un, ["Type", "Key", "Detail"], uniq)

    order = ["Instructions", "Config", "Uploader", "Extract", "GL Master", "Control", "Unmapped"]
    for target_idx, name in enumerate(order):
        if name in wb.sheetnames:
            current = wb.sheetnames.index(name)
            if current != target_idx:
                wb.move_sheet(name, offset=target_idx - current)

    wb.save(output_path)
    print(f"Saved {output_path} with {len(builder.rows)} uploader lines, {len(uniq)} unmapped keys")

    # Print control summary
    by_v = defaultdict(lambda: [0.0, 0.0])
    for row in builder.rows:
        by_v[row["Voucher No"]][0] += row["Debit Amount"] or 0
        by_v[row["Voucher No"]][1] += row["Credit Amount"] or 0
    print("\nVoucher balance check:")
    for v in sorted(by_v):
        d, c = by_v[v]
        print(f"  V{v}: Dr={d:,.2f} Cr={c:,.2f} Diff={d-c:,.2f}")


def main():
    ap = argparse.ArgumentParser(description="Build RetailX Uploader from Extract")
    ap.add_argument("--input", default="RetailX.xlsx")
    ap.add_argument("--output", default="RetailX.xlsx")
    args = ap.parse_args()
    build(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
