#!/usr/bin/env python3
"""
MP Summary Uploader Builder
---------------------------
Builds the Uploader Format sheet from Summary + GL Backup.

Rules implemented:
  1. Prepaid / postpaid expense entries at state level (state_code_to)
  2. TCS / TDS entries use tcs_state_code_to for state
  3. GL codes looked up from GL Backup by Summary header
  4. IGST input: IN-DL -> 142067, all other states -> 142013
  5. Debtor and provision ledgers always posted at IN-OTH
  6. Prior-month MEC-FKMP OI report drives OI opening (expense reversal) vouchers
  7. Negative Summary values -> Debit expense/GST input, Credit debtor (reversed for OI opening)

Usage:
  python3 build_mp_uploader.py
  python3 build_mp_uploader.py --input "MP Summary.xlsx" --output "MP Summary.xlsx"
"""

from __future__ import annotations

import argparse
import calendar
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

OI_FILE = "MEC-FKMP-OPEN-INVOICE-FLOW.csv"
CREDITOR_FILE = "FKMP-CREDITOR-REPORT (1).csv"

IGST_GL_DL = 142067
IGST_GL_OTHER = 142013

TCS_COLUMNS = {
    "sum(invoice_tcs_cgst_amount)": "sum(invoice_tcs_cgst_amount)",
    "sum(invoice_tcs_sgst_amount)": "sum(invoice_tcs_sgst_amount)",
    "sum(invoice_tcs_igst_amount)": "sum(invoice_tcs_igst_amount)",
}
TDS_COLUMN = "sum(invoice_tds_income_tax_amount)"

GST_COLUMNS = {
    "sum(sgst_utgst_total_amount)",
    "sum(cgst_total_amount)",
    "sum(igst_total_amount)",
}

SKIP_COLUMNS = {"sum(due_amount)"}

DEBTOR_KEYS = {
    "prepaid": "Prepaid Debtor",
    "postpaid": "Postpaid Debtor",
    "provision": "Provision Ledger",
    "vd": "VD Debtor",
}

DEFAULT_VOUCHERS = [
    {
        "voucher_no": 28,
        "key": "OI_CLOSING_PREPAID",
        "month_scope": "current",
        "filename": OI_FILE,
        "order_type": "prepaid",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "prepaid",
        "narration": "MP_Charges_OI closing-Prepaid for the month of {month}",
        "provision_reverse": "No",
    },
    {
        "voucher_no": 29,
        "key": "OI_CLOSING_POSTPAID",
        "month_scope": "current",
        "filename": OI_FILE,
        "order_type": "postpaid",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "postpaid",
        "narration": "MP_Charges_OI closing-Postpaid for the month of {month}",
        "provision_reverse": "No",
    },
    {
        "voucher_no": 30,
        "key": "CREDITOR_POSTPAID",
        "month_scope": "current",
        "filename": CREDITOR_FILE,
        "order_type": "postpaid",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "postpaid",
        "narration": "MP Charges creditor for the month of {month} - Postpaid",
        "provision_reverse": "No",
    },
    {
        "voucher_no": 31,
        "key": "CREDITOR_PREPAID",
        "month_scope": "current",
        "filename": CREDITOR_FILE,
        "order_type": "prepaid",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "prepaid",
        "narration": "MP Charges creditor for the month of {month} - Prepaid",
        "provision_reverse": "No",
    },
    {
        "voucher_no": 32,
        "key": "OI_OPENING_PREPAID",
        "month_scope": "prior",
        "filename": OI_FILE,
        "order_type": "prepaid",
        "reverse": True,
        "voucher_type": "AR-Journal",
        "debtor_scope": "prepaid",
        "narration": "MP_Charges_OI opening-Prepaid for the month of {month}",
        "provision_reverse": "No",
    },
    {
        "voucher_no": 33,
        "key": "OI_OPENING_POSTPAID",
        "month_scope": "prior",
        "filename": OI_FILE,
        "order_type": "postpaid",
        "reverse": True,
        "voucher_type": "AR-Journal",
        "debtor_scope": "postpaid",
        "narration": "MP_Charges_OI opening-postpaid for the month of {month}",
        "provision_reverse": "No",
    },
    {
        "voucher_no": 34,
        "key": "PROVISION",
        "month_scope": "current",
        "filename": "Provision",
        "order_type": "Provision",
        "reverse": False,
        "voucher_type": "AR-Provision",
        "debtor_scope": "provision",
        "narration": "MP charges Provision - subsequent return & undelivered for the month of {month}",
        "provision_reverse": "Yes",
    },
    {
        "voucher_no": 35,
        "key": "VD",
        "month_scope": "current",
        "filename": "VD",
        "order_type": "VD",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "vd",
        "narration": "MP  charges volume discount for the month of {month}",
        "provision_reverse": "No",
    },
    {
        "voucher_no": 36,
        "key": "PBO_VD",
        "month_scope": "current",
        "filename": "PBO VD",
        "order_type": "VD",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "vd",
        "narration": "MP charges volume discount on PBO adjustment for the month  of {month}",
        "provision_reverse": "No",
    },
    {
        "voucher_no": 41,
        "key": "TCS_POSTPAID",
        "month_scope": "both",
        "order_type": "postpaid",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "postpaid",
        "narration": "TCS GST Receivable ( Postpaid) For the month of {month}",
        "provision_reverse": "No",
        "tcs_tds": "tcs",
    },
    {
        "voucher_no": 42,
        "key": "TDS_POSTPAID",
        "month_scope": "both",
        "order_type": "postpaid",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "postpaid",
        "narration": "TDS Receivable ( Postpaid) For the month of {month}",
        "provision_reverse": "No",
        "tcs_tds": "tds",
    },
    {
        "voucher_no": 43,
        "key": "TCS_PREPAID",
        "month_scope": "both",
        "order_type": "prepaid",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "prepaid",
        "narration": "TCS GST Receivable ( Prepaid) For the month of {month}",
        "provision_reverse": "No",
        "tcs_tds": "tcs",
    },
    {
        "voucher_no": 44,
        "key": "TDS_PREPAID",
        "month_scope": "both",
        "order_type": "prepaid",
        "reverse": False,
        "voucher_type": "AR-Journal",
        "debtor_scope": "prepaid",
        "narration": "TDS Receivable ( Prepaid) For the month of {month}",
        "provision_reverse": "No",
        "tcs_tds": "tds",
    },
]

UPLOADER_HEADERS = [
    "VoucherType",
    "Account Name",
    "Date",
    "Ref New Field",
    "Ledger Narration",
    "Voucher No",
    " State Name",
    "Function",
    "Location",
    "Debit Amount",
    "Credit Amount",
    "Narration",
    "Sl no",
    "Invoice  Type",
    "Company Code",
    "Is provision reverse",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def r2(v: float) -> float:
    return round(float(v or 0.0), 2)


def norm_header(h) -> str:
    return str(h or "").strip().replace("\n", "")


def normalize_state(state) -> str | None:
    if state is None:
        return None
    s = str(state).strip().upper()
    if not s or s in {"NA", "NONE", "N/A"}:
        return None
    if s.startswith("IN-"):
        return s
    if len(s) == 2:
        return f"IN-{s}"
    return s


def tcs_tds_state(row: dict) -> str | None:
    """Use tcs_state_code_to; fall back to state_code_to when TCS state is NA."""
    return normalize_state(row.get("tcs_state_code_to")) or normalize_state(
        row.get("state_code_to")
    )


def normalize_month_label(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%b'%y")
    text = str(value).strip().replace("\u2019", "'").replace("'", "'")
    if not text:
        return None
    try:
        year, month, _ = parse_month_label(text)
        return datetime(year, month, 1).strftime("%b'%y")
    except ValueError:
        return text


def parse_month_label(label: str) -> tuple[int, int, str]:
    text = str(label).strip().replace("'", "").replace("'", "")
    for fmt in ("%b%y", "%b%Y"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.year, dt.month, label
        except ValueError:
            continue
    raise ValueError(f"Cannot parse month label: {label!r}")


def prior_month_label(label: str) -> str:
    year, month, _ = parse_month_label(normalize_month_label(label) or label)
    if month == 1:
        year -= 1
        month = 12
    else:
        month -= 1
    return datetime(year, month, 1).strftime("%b'%y")


def tcs_tds_month_labels(rows: list[dict], current_month: str) -> list[str]:
    """Use the latest two month labels present in Summary for TCS/TDS vouchers."""
    labels: list[str] = []
    seen: set[str] = set()
    for row in rows:
        lbl = normalize_month_label(row.get("Month"))
        if lbl and lbl not in seen:
            seen.add(lbl)
            labels.append(lbl)
    labels.sort(key=lambda m: parse_month_label(m)[:2])
    if len(labels) >= 2:
        return labels[-2:]
    if labels:
        return [labels[-1]]
    current = normalize_month_label(current_month) or current_month
    return [current, prior_month_label(current)]


def month_end_date(label: str) -> datetime:
    year, month, _ = parse_month_label(label)
    last = calendar.monthrange(year, month)[1]
    return datetime(year, month, last)


def sign_to_dr_cr(amount: float, reverse: bool = False) -> tuple[float, float]:
    """Negative -> debit GL; positive -> credit GL. Reverse flips for OI opening."""
    if abs(amount) < 0.005:
        return 0.0, 0.0
    if amount < 0:
        dr, cr = abs(amount), 0.0
    else:
        dr, cr = 0.0, abs(amount)
    if reverse:
        dr, cr = cr, dr
    return dr, cr


def tcs_tds_to_dr_cr(amount: float) -> tuple[float, float]:
    """After state-level netting: negative -> debit receivable, positive -> credit."""
    return sign_to_dr_cr(amount, reverse=False)


def is_expense_column(header: str, all_headers: set[str]) -> bool:
    h = norm_header(header)
    if h in SKIP_COLUMNS:
        return False
    if h in GST_COLUMNS or h in TCS_COLUMNS or h == TDS_COLUMN:
        return False
    if not h.startswith("sum(") and not h.startswith("SUM("):
        return False
    return True


# ---------------------------------------------------------------------------
# Readers
# ---------------------------------------------------------------------------


def read_gl_backup(ws) -> dict[str, int | str]:
    mapping: dict[str, int | str] = {}
    for r in range(2, ws.max_row + 1):
        field = norm_header(ws.cell(r, 1).value)
        gl = ws.cell(r, 2).value
        if not field:
            continue
        mapping[field] = gl
        mapping[field.lower()] = gl
    return mapping


def to_amount(val) -> float | None:
    if val in (None, ""):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if s.startswith("="):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def read_summary(ws) -> tuple[list[str], list[dict]]:
    headers = [norm_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[i - 1]: ws.cell(r, i).value for i in range(1, len(headers) + 1)}
        if row.get("Month"):
            rows.append(row)
    return headers, rows


def read_config(wb) -> dict:
    cfg = {
        "month": None,
        "company_code": "HRFK",
        "invoice_type": "B2B",
        "location": "Business Operations",
        "function": "Others",
        "vouchers": {v["key"]: v["voucher_no"] for v in DEFAULT_VOUCHERS},
    }
    if "Config" not in wb.sheetnames:
        return cfg
    ws = wb["Config"]
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 1).value
        val = ws.cell(r, 2).value
        if not key:
            continue
        k = str(key).strip().lower()
        if k == "month":
            cfg["month"] = str(val).strip() if val else None
        elif k == "company code":
            cfg["company_code"] = str(val).strip()
        elif k == "invoice type":
            cfg["invoice_type"] = str(val).strip()
        elif k == "location":
            cfg["location"] = str(val).strip()
        elif k == "function":
            cfg["function"] = str(val).strip()
        elif k.startswith("voucher "):
            vkey = k.replace("voucher ", "").strip().upper()
            for spec in DEFAULT_VOUCHERS:
                if spec["key"] == vkey and val is not None:
                    cfg["vouchers"][vkey] = int(val)
    return cfg


def detect_month(rows: list[dict], cfg_month: str | None) -> str:
    if cfg_month:
        return cfg_month
    labels = sorted(
        {str(r["Month"]).strip() for r in rows if r.get("Month")},
        key=lambda m: parse_month_label(m)[:2],
    )
    return labels[-1] if labels else "Unknown"


def gl_for_column(gl_map: dict, header: str, state: str) -> int | None:
    h = norm_header(header)
    if h == "sum(igst_total_amount)":
        return IGST_GL_DL if state == "IN-DL" else IGST_GL_OTHER
    raw = gl_map.get(h) or gl_map.get(h.lower())
    if raw in (None, "", "NA"):
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def debtor_gl(gl_map: dict, scope: str) -> int | None:
    key = DEBTOR_KEYS.get(scope)
    if not key:
        return None
    raw = gl_map.get(key)
    if raw in (None, "", "NA"):
        return None
    return int(raw)


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------


class MPUploaderBuilder:
    def __init__(self, company_code, invoice_type, location, function):
        self.company_code = company_code
        self.invoice_type = invoice_type
        self.location = location
        self.function = function
        self.rows: list[dict] = []
        self.unmapped: list[dict] = []
        self.sl = 0

    def add(
        self,
        *,
        voucher_no,
        account,
        date,
        state,
        debit,
        credit,
        narration,
        voucher_type="AR-Journal",
        provision_reverse="No",
    ):
        debit, credit = r2(debit), r2(credit)
        if abs(debit) < 0.005 and abs(credit) < 0.005:
            return
        if debit < 0 and credit == 0:
            credit, debit = -debit, 0.0
        if credit < 0 and debit == 0:
            debit, credit = -credit, 0.0
        self.sl += 1
        self.rows.append(
            {
                "VoucherType": voucher_type,
                "Account Name": account,
                "Date": date,
                "Ref New Field": None,
                "Ledger Narration": None,
                "Voucher No": voucher_no,
                " State Name": state,
                "Function": self.function,
                "Location": self.location,
                "Debit Amount": debit,
                "Credit Amount": credit,
                "Narration": narration,
                "Sl no": self.sl,
                "Invoice  Type": self.invoice_type,
                "Company Code": self.company_code,
                "Is provision reverse": provision_reverse,
            }
        )

    def note_unmapped(self, kind, key, detail=""):
        self.unmapped.append({"Type": kind, "Key": key, "Detail": detail})


def filter_tcs_tds_rows(
    rows: list[dict],
    months: list[str],
    order_type: str,
) -> list[dict]:
    month_set = {normalize_month_label(m) or str(m).strip() for m in months}
    out = []
    for row in rows:
        row_month = normalize_month_label(row.get("Month"))
        if row_month not in month_set:
            continue
        if str(row.get("order_type", "")).strip() != order_type:
            continue
        out.append(row)
    return out


def filter_rows(
    rows: list[dict],
    month: str,
    order_type: str,
    filename: str | None = None,
) -> list[dict]:
    out = []
    for row in rows:
        if str(row.get("Month", "")).strip() != month:
            continue
        if str(row.get("order_type", "")).strip() != order_type:
            continue
        if filename is not None and str(row.get("filename", "")).strip() != filename:
            continue
        out.append(row)
    return out


def aggregate_expense_lines(
    rows: list[dict],
    headers: list[str],
    gl_map: dict,
    reverse: bool,
) -> list[tuple[str, str, float, int]]:
    """Return list of (state, column_header, amount, gl)."""
    agg: dict[tuple[str, str], float] = defaultdict(float)
    header_set = set(headers)
    expense_headers = [h for h in headers if is_expense_column(h, header_set)]
    gst_headers = [h for h in headers if norm_header(h) in GST_COLUMNS]

    for row in rows:
        state = normalize_state(row.get("state_code_to"))
        if not state:
            continue
        for col in expense_headers + gst_headers:
            amount = to_amount(row.get(col))
            if amount is None or abs(amount) < 0.005:
                continue
            gl = gl_for_column(gl_map, col, state)
            if gl is None:
                continue
            agg[(state, col)] += amount

    lines = []
    for (state, col), amount in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        gl = gl_for_column(gl_map, col, state)
        if gl is None:
            continue
        lines.append((state, col, amount, gl))
    return lines


def gen_expense_voucher(
    builder: MPUploaderBuilder,
    rows: list[dict],
    headers: list[str],
    gl_map: dict,
    spec: dict,
    month_label: str,
    book_date: datetime,
    voucher_no: int,
):
    reverse = spec.get("reverse", False)
    narration = spec["narration"].format(month=month_label)
    lines = aggregate_expense_lines(rows, headers, gl_map, reverse)

    total_dr = 0.0
    total_cr = 0.0
    for state, col, amount, gl in lines:
        dr, cr = sign_to_dr_cr(amount, reverse=reverse)
        if gl is None:
            builder.note_unmapped("GL", col, f"state={state}")
            continue
        builder.add(
            voucher_no=voucher_no,
            account=gl,
            date=book_date,
            state=state,
            debit=dr,
            credit=cr,
            narration=narration,
            voucher_type=spec["voucher_type"],
            provision_reverse=spec["provision_reverse"],
        )
        total_dr += dr
        total_cr += cr

    d_gl = debtor_gl(gl_map, spec["debtor_scope"])
    if d_gl is None:
        builder.note_unmapped("Debtor", spec["debtor_scope"])
        return

    # Balance debtor at IN-OTH
    diff = r2(total_dr - total_cr)
    if diff > 0:
        builder.add(
            voucher_no=voucher_no,
            account=d_gl,
            date=book_date,
            state="IN-OTH",
            debit=0,
            credit=diff,
            narration=narration,
            voucher_type=spec["voucher_type"],
            provision_reverse=spec["provision_reverse"],
        )
    elif diff < 0:
        builder.add(
            voucher_no=voucher_no,
            account=d_gl,
            date=book_date,
            state="IN-OTH",
            debit=-diff,
            credit=0,
            narration=narration,
            voucher_type=spec["voucher_type"],
            provision_reverse=spec["provision_reverse"],
        )


def gen_tcs_tds_voucher(
    builder: MPUploaderBuilder,
    rows: list[dict],
    headers: list[str],
    gl_map: dict,
    spec: dict,
    month_label: str,
    book_date: datetime,
    voucher_no: int,
):
    narration = spec["narration"].format(month=month_label)
    mode = spec["tcs_tds"]
    cols = list(TCS_COLUMNS.keys()) if mode == "tcs" else [TDS_COLUMN]

    # Net all Summary rows at state level (per GL) before posting.
    agg: dict[tuple[str, int], float] = defaultdict(float)
    for row in rows:
        state = tcs_tds_state(row)
        if not state:
            continue
        for col in cols:
            amount = to_amount(row.get(col))
            if amount is None or abs(amount) < 0.005:
                continue
            gl = gl_for_column(gl_map, col, state)
            if gl is None:
                continue
            agg[(state, gl)] += amount

    total_dr = 0.0
    total_cr = 0.0
    for (state, gl), amount in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1])):
        if abs(amount) < 0.005:
            continue
        dr, cr = tcs_tds_to_dr_cr(amount)
        builder.add(
            voucher_no=voucher_no,
            account=gl,
            date=book_date,
            state=state,
            debit=dr,
            credit=cr,
            narration=narration,
            voucher_type=spec["voucher_type"],
            provision_reverse=spec["provision_reverse"],
        )
        total_dr += dr
        total_cr += cr

    d_gl = debtor_gl(gl_map, spec["debtor_scope"])
    if d_gl is None:
        builder.note_unmapped("Debtor", spec["debtor_scope"])
        return

    diff = r2(total_dr - total_cr)
    if diff > 0:
        builder.add(
            voucher_no=voucher_no,
            account=d_gl,
            date=book_date,
            state="IN-OTH",
            debit=0,
            credit=diff,
            narration=narration,
            voucher_type=spec["voucher_type"],
            provision_reverse=spec["provision_reverse"],
        )
    elif diff < 0:
        builder.add(
            voucher_no=voucher_no,
            account=d_gl,
            date=book_date,
            state="IN-OTH",
            debit=-diff,
            credit=0,
            narration=narration,
            voucher_type=spec["voucher_type"],
            provision_reverse=spec["provision_reverse"],
        )


def balance_voucher_rows(rows: list[dict], tol: float = 0.05):
    by_v: dict[int, list[int]] = defaultdict(list)
    for i, row in enumerate(rows):
        by_v[row["Voucher No"]].append(i)
    for idxs in by_v.values():
        deb = sum(rows[i]["Debit Amount"] or 0 for i in idxs)
        cred = sum(rows[i]["Credit Amount"] or 0 for i in idxs)
        diff = r2(deb - cred)
        if abs(diff) < 0.005 or abs(diff) > 50:
            continue
        candidates = [
            i
            for i in idxs
            if rows[i][" State Name"] == "IN-OTH"
            and ((rows[i]["Debit Amount"] or 0) or (rows[i]["Credit Amount"] or 0))
        ] or idxs
        if diff > 0:
            i = max(candidates, key=lambda x: rows[x]["Credit Amount"] or 0)
            if (rows[i]["Credit Amount"] or 0) > 0:
                rows[i]["Credit Amount"] = r2(rows[i]["Credit Amount"] + diff)
            else:
                i = max(candidates, key=lambda x: rows[x]["Debit Amount"] or 0)
                rows[i]["Debit Amount"] = r2(rows[i]["Debit Amount"] - diff)
        else:
            i = max(candidates, key=lambda x: rows[x]["Debit Amount"] or 0)
            if (rows[i]["Debit Amount"] or 0) > 0:
                rows[i]["Debit Amount"] = r2(rows[i]["Debit Amount"] - diff)
            else:
                i = max(candidates, key=lambda x: rows[x]["Credit Amount"] or 0)
                rows[i]["Credit Amount"] = r2(rows[i]["Credit Amount"] + (-diff))


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


def ensure_sheet(wb, name: str):
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    return wb[name]


def write_config(ws, month_label: str, book_date: datetime, voucher_map: dict):
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = "Setting"
    ws["B1"] = "Value"
    ws["A1"].font = HEADER_FONT
    ws["B1"].font = HEADER_FONT
    ws["A1"].fill = HEADER_FILL
    ws["B1"].fill = HEADER_FILL
    rows = [
        ("Month", month_label),
        ("Booking Date", book_date),
        ("Company Code", "HRFK"),
        ("Invoice Type", "B2B"),
        ("Location", "Business Operations"),
        ("Function", "Others"),
    ]
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)
        if k == "Booking Date" and isinstance(v, datetime):
            ws.cell(i, 2).number_format = "DD-MMM-YYYY"
    r = len(rows) + 2
    ws.cell(r, 1, "Voucher Key")
    ws.cell(r, 2, "Voucher No")
    ws.cell(r, 1).font = Font(bold=True)
    ws.cell(r, 2).font = Font(bold=True)
    for j, spec in enumerate(DEFAULT_VOUCHERS, r + 1):
        ws.cell(j, 1, spec["key"])
        ws.cell(j, 2, voucher_map.get(spec["key"], spec["voucher_no"]))


def write_control(ws, builder_rows: list[dict]):
    ws.delete_rows(1, ws.max_row)
    headers = ["Voucher No", "Narration", "Debit", "Credit", "Difference"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    by_v: dict[int, dict] = defaultdict(lambda: {"d": 0.0, "c": 0.0, "n": ""})
    for row in builder_rows:
        v = row["Voucher No"]
        by_v[v]["d"] += row["Debit Amount"] or 0
        by_v[v]["c"] += row["Credit Amount"] or 0
        by_v[v]["n"] = row["Narration"]
    for i, v in enumerate(sorted(by_v), 2):
        ws.cell(i, 1, v)
        ws.cell(i, 2, by_v[v]["n"])
        ws.cell(i, 3, r2(by_v[v]["d"])).number_format = "#,##0.00"
        ws.cell(i, 4, r2(by_v[v]["c"])).number_format = "#,##0.00"
        diff = r2(by_v[v]["d"] - by_v[v]["c"])
        cell = ws.cell(i, 5, diff)
        cell.number_format = "#,##0.00"
        if abs(diff) > 1:
            cell.font = Font(color="9C0006", bold=True)


def write_unmapped(ws, items: list[dict]):
    ws.delete_rows(1, ws.max_row)
    for c, h in enumerate(["Type", "Key", "Detail"], 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    seen = set()
    r = 2
    for item in items:
        key = (item["Type"], item["Key"])
        if key in seen:
            continue
        seen.add(key)
        ws.cell(r, 1, item["Type"])
        ws.cell(r, 2, item["Key"])
        ws.cell(r, 3, item.get("Detail", ""))
        r += 1


def write_instructions(ws):
    ws.delete_rows(1, ws.max_row)
    lines = [
        ("MP Summary Dynamic Uploader", True),
        ("", False),
        ("How to refresh", True),
        ("1) Paste latest Summary pivot data (keep header row).", False),
        ("2) Maintain GL Backup mappings for every Summary fee column.", False),
        ("3) Run: python3 build_mp_uploader.py", False),
        ("   Or import GenerateMPUploader.bas and run the macro (Windows/Mac Excel).", False),
        ("", False),
        ("Voucher series", True),
        ("28-29  OI closing (current month MEC OI) — Prepaid / Postpaid", False),
        ("30-31  Creditor (current month FKMP creditor report) — Postpaid / Prepaid", False),
        ("32-33  OI opening reversal (prior month MEC OI) — Prepaid / Postpaid", False),
        ("34     Provision (order_type = Provision)", False),
        ("35-36  Volume discount (VD / PBO VD rows)", False),
        ("41-44  TCS / TDS receivable by prepaid & postpaid (current + prior month)", False),
        ("", False),
        ("Posting rules", True),
        ("Expense & GST lines use state_code_to at state level.", False),
        ("TCS / TDS lines use tcs_state_code_to (fallback state_code_to), net at state level per GL, and include the latest two Summary months.", False),
        ("TCS/TDS: negative net -> debit receivable; positive net -> credit receivable.", False),
        ("IGST input: IN-DL -> 142067; other states -> 142013.", False),
        ("Debtor & provision ledgers always use IN-OTH.", False),
        ("Negative Summary values -> Debit expense/GST, Credit debtor.", False),
        ("OI opening vouchers reverse the prior-month OI expense entries.", False),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        ws.cell(i, 1, text)
        if bold:
            ws.cell(i, 1).font = Font(bold=True, color="1F4E79", size=12 if i == 1 else 11)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 110


# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------


def build(input_path: Path, output_path: Path):
    wb = load_workbook(input_path)
    if "Summary" not in wb.sheetnames or "GL Backup" not in wb.sheetnames:
        raise SystemExit("Workbook must contain Summary and GL Backup sheets")

    # Cached values for formula cells in Summary
    wb_vals = load_workbook(input_path, data_only=True)
    headers, summary_rows = read_summary(wb_vals["Summary"])
    gl_map = read_gl_backup(wb["GL Backup"])
    cfg = read_config(wb)

    month_label = detect_month(summary_rows, cfg["month"])
    book_date = month_end_date(month_label)
    prior_label = prior_month_label(month_label)
    tcs_tds_months = tcs_tds_month_labels(summary_rows, month_label)

    builder = MPUploaderBuilder(
        company_code=cfg["company_code"],
        invoice_type=cfg["invoice_type"],
        location=cfg["location"],
        function=cfg["function"],
    )

    print(f"Month: {month_label} | Prior OI month: {prior_label} | TCS/TDS months: {tcs_tds_months} | Date: {book_date.date()}")
    print(f"Summary rows: {len(summary_rows)} | GL mappings: {len(gl_map)}")

    for spec in DEFAULT_VOUCHERS:
        vno = cfg["vouchers"].get(spec["key"], spec["voucher_no"])
        month = month_label if spec["month_scope"] == "current" else prior_label

        if spec.get("tcs_tds"):
            scoped = filter_tcs_tds_rows(summary_rows, tcs_tds_months, spec["order_type"])
            if not scoped:
                print(
                    f"  skip voucher {vno} {spec['key']} — no Summary rows for "
                    f"{' / '.join(tcs_tds_months)}"
                )
                continue
            before = len(builder.rows)
            gen_tcs_tds_voucher(
                builder, scoped, headers, gl_map, spec, month_label, book_date, vno
            )
            print(f"  voucher {vno} {spec['key']}: {len(builder.rows) - before} lines")
            continue

        filename = spec.get("filename")
        scoped = filter_rows(summary_rows, month, spec["order_type"], filename)
        if not scoped:
            print(f"  skip voucher {vno} {spec['key']} — no Summary rows for {month}")
            continue
        before = len(builder.rows)
        gen_expense_voucher(
            builder, scoped, headers, gl_map, spec, month_label, book_date, vno
        )
        print(f"  voucher {vno} {spec['key']}: {len(builder.rows) - before} lines")

    balance_voucher_rows(builder.rows)

    ws_up = ensure_sheet(wb, "Uploader Format")
    if ws_up.max_row >= 1:
        ws_up.delete_rows(1, ws_up.max_row)
    for c, h in enumerate(UPLOADER_HEADERS, 1):
        cell = ws_up.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for i, row in enumerate(builder.rows, 2):
        for c, h in enumerate(UPLOADER_HEADERS, 1):
            val = row.get(h)
            cell = ws_up.cell(i, c, val)
            cell.border = THIN
            if h == "Date" and isinstance(val, datetime):
                cell.number_format = "DD-MMM-YYYY"
            if h in ("Debit Amount", "Credit Amount"):
                cell.number_format = "#,##0.00"

    widths = [12, 14, 12, 12, 14, 12, 12, 10, 18, 14, 14, 50, 8, 12, 12, 16]
    for i, w in enumerate(widths, 1):
        ws_up.column_dimensions[get_column_letter(i)].width = w

    write_config(ensure_sheet(wb, "Config"), month_label, book_date, cfg["vouchers"])
    write_instructions(ensure_sheet(wb, "Instructions"))
    write_control(ensure_sheet(wb, "Control"), builder.rows)
    write_unmapped(ensure_sheet(wb, "Unmapped"), builder.unmapped)

    order = [
        "Instructions",
        "Config",
        "Uploader Format",
        "Summary",
        "GL Backup",
        "Control",
        "Unmapped",
    ]
    for target_idx, name in enumerate(order):
        if name in wb.sheetnames:
            current = wb.sheetnames.index(name)
            if current != target_idx:
                wb.move_sheet(name, offset=target_idx - current)

    wb.save(output_path)
    print(f"Saved {output_path} with {len(builder.rows)} uploader lines")

    by_v: dict[int, list[float]] = defaultdict(lambda: [0.0, 0.0])
    for row in builder.rows:
        by_v[row["Voucher No"]][0] += row["Debit Amount"] or 0
        by_v[row["Voucher No"]][1] += row["Credit Amount"] or 0
    print("\nVoucher balance check:")
    for v in sorted(by_v):
        d, c = by_v[v]
        status = "OK" if abs(d - c) < 1 else f"DIFF {d - c:,.2f}"
        print(f"  V{v}: Dr={d:,.2f} Cr={c:,.2f} -> {status}")


def main():
    ap = argparse.ArgumentParser(description="Build MP Summary Uploader from Summary sheet")
    ap.add_argument("--input", default="MP Summary.xlsx")
    ap.add_argument("--output", default="MP Summary.xlsx")
    args = ap.parse_args()
    build(Path(args.input), Path(args.output))


if __name__ == "__main__":
    main()
