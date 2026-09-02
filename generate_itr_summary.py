#!/usr/bin/env python3
"""Generate combined ITR filing summary Excel from YES Bank XLS and SCB PDF statements."""

import re
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pypdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
YES_STATEMENT_PATH = Path("/workspace/Account_Statement_01_Apr_2025-31_Mar_2026.xls")
SCB_PDF_PATH = Path("/workspace/ilovepdf_merged.pdf")
OUTPUT_PATH = Path("/workspace/Arun/ITR_Summary_AY_2026-27.xlsx")

ASSESSEE = {
    "name": "ARUN SHANKAR AWASTHI / MR ARUN AWASTHI",
    "email": "arun15web@gmail.com",
    "phone": "918882283917",
    "financial_year": "FY 2025-26",
    "assessment_year": "AY 2026-27",
    "statement_period": "01/04/2025 to 31/03/2026",
}

ACCOUNTS = [
    {
        "bank": "YES Bank",
        "account_no": "014091900000507",
        "account_type": "Savings - Smart Salary Platinum",
        "branch": "Badshapur, Haryana",
        "ifsc": "YESB0000140",
        "source": "Excel (.xls)",
    },
    {
        "bank": "Standard Chartered Bank",
        "account_no": "54410610545",
        "account_type": "Supervalue Savings",
        "branch": "New Friends Colony, New Delhi",
        "ifsc": "SCBL0036034",
        "source": "PDF (merged statements)",
    },
]

YES_ACCOUNT = "014091900000507"
SCB_ACCOUNT = "54410610545"
YES_IFSC = "YESB0000140"
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
LABEL_FONT = Font(bold=True)
MONEY_FMT = "#,##0.00"
DATE_FMT = "DD-MMM-YYYY"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SALARY_EXCLUDE_KEYWORDS = ("REFUND", "REVERT", "CREDIT CARD", "CREDITCARD")
SALARY_INCLUDE_KEYWORDS = (
    "SALARY",
    "PAY ARUN",
    "ARUN PAYMENT",
    "PAY SALARY",
    "ARUNAWAS",
    "PATYCASH",
    "URGENT//",
    "/SALARY",
)


def parse_amount(value: str) -> float:
    if not value:
        return 0.0
    return float(str(value).replace(",", ""))


def parse_date(value: str) -> datetime | None:
    for fmt in ("%d %b %Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(value.strip(), fmt)
        except ValueError:
            continue
    return None


def normalize_description(description: str) -> str:
    text = re.sub(r"\s+", " ", description.replace("\n", " ")).strip()
    for marker in (" Total ", "REWARD POINTS", "Dear Client", "STATEMENT DATE", "Date Value Date"):
        if marker in text:
            text = text.split(marker, 1)[0].strip()
    return text


def extract_trailing_amounts(text: str) -> tuple[str, list[float]]:
    amount_pattern = r"([\d,]+\.\d{2})"
    amounts = re.findall(amount_pattern, text)
    description = re.sub(
        r"\s+" + amount_pattern + r"(?:\s+" + amount_pattern + r")?\s*$",
        "",
        text,
        flags=re.DOTALL,
    ).strip()
    return description, [parse_amount(value) for value in amounts]


def get_row_values(row) -> dict[int, str]:
    cells = row.findall("ss:Cell", NS)
    result: dict[int, str] = {}
    col = 0
    for cell in cells:
        idx = cell.get("{urn:schemas-microsoft-com:office:spreadsheet}Index")
        if idx:
            col = int(idx) - 1
        data = cell.find("ss:Data", NS)
        val = (data.text or "").strip() if data is not None else ""
        result[col] = val
        col += 1
    return result


def load_yes_bank_transactions() -> list[dict]:
    root = ET.fromstring(YES_STATEMENT_PATH.read_text(encoding="utf-8"))
    rows = root.findall(".//ss:Worksheet/ss:Table/ss:Row", NS)
    header_idx = next(
        i for i, row in enumerate(rows) if get_row_values(row).get(0) == "Transaction Date"
    )

    transactions = []
    for row in rows[header_idx + 1 :]:
        vals = get_row_values(row)
        if not re.match(r"\d{2} \w{3} \d{4}", vals.get(0, "")):
            continue
        transactions.append(
            {
                "bank": "YES Bank",
                "account_no": "014091900000507",
                "txn_date": parse_date(vals.get(0, "")),
                "value_date": parse_date(vals.get(1, "")),
                "ref_no": vals.get(2, ""),
                "description": normalize_description(vals.get(3, "")),
                "withdrawal": parse_amount(vals.get(4, "")),
                "deposit": parse_amount(vals.get(5, "")),
                "balance": parse_amount(vals.get(6, "")),
            }
        )
    return transactions


def load_scb_pdf_transactions() -> list[dict]:
    reader = pypdf.PdfReader(str(SCB_PDF_PATH))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    text = re.sub(r"(\d{2} \w{3})\s*\n\s*(\d{4})", r"\1 \2", text)
    text = re.sub(r"Page\s+of\s+\d+\s+\d+", "", text)
    text = re.sub(r"MR ARUN AWASTHI\s*", "", text)

    txn_start = re.compile(r"(?=\d{2} \w{3} \d{4}\s+\d{2} \w{3} \d{4}\s+)")
    date_pair = re.compile(r"^(\d{2} \w{3} \d{4})\s+(\d{2} \w{3} \d{4})\s+(.*)$", re.DOTALL)

    parsed: list[dict] = []
    previous_balance: float | None = None

    for block in txn_start.split(text):
        match = date_pair.match(block.strip())
        if not match:
            continue

        txn_date_s, value_date_s, rest = match.groups()
        rest = normalize_description(rest.strip())
        description, amounts = extract_trailing_amounts(rest)

        if not description or description.upper().startswith("TOTAL"):
            continue

        if description.startswith("BALANCE FORWARD") and amounts:
            previous_balance = amounts[-1]
            continue

        if not amounts:
            continue

        ref_no = ""
        if len(amounts) == 1:
            balance = amounts[0]
            deposit = withdrawal = 0.0
        else:
            amount = amounts[-2]
            balance = amounts[-1]
            deposit = withdrawal = 0.0
            if previous_balance is not None:
                delta = round(balance - previous_balance, 2)
                if delta > 0:
                    deposit = amount
                elif delta < 0:
                    withdrawal = amount

        previous_balance = balance

        if description.startswith("IMPS "):
            ref_match = re.match(r"IMPS\s+(\S+)", description)
            ref_no = ref_match.group(1) if ref_match else ""
        elif description.startswith("NEFT "):
            ref_match = re.match(r"NEFT\s+(\S+)", description)
            ref_no = ref_match.group(1) if ref_match else ""
        elif description.startswith("UPI/"):
            ref_match = re.match(r"UPI/(\S+)", description)
            ref_no = ref_match.group(1) if ref_match else ""

        parsed.append(
            {
                "bank": "Standard Chartered Bank",
                "account_no": "54410610545",
                "txn_date": parse_date(txn_date_s),
                "value_date": parse_date(value_date_s),
                "ref_no": ref_no,
                "description": description,
                "withdrawal": withdrawal,
                "deposit": deposit,
                "balance": balance,
            }
        )

    return parsed


def load_all_transactions() -> list[dict]:
    yes_txns = load_yes_bank_transactions()
    scb_txns = load_scb_pdf_transactions()
    return yes_txns + scb_txns


def is_infutive_salary(txn: dict) -> bool:
    desc = txn["description"].upper()
    if "INFUTIVE" not in desc or txn["deposit"] <= 0:
        return False
    if any(keyword in desc for keyword in SALARY_EXCLUDE_KEYWORDS):
        return False
    return any(keyword in desc for keyword in SALARY_INCLUDE_KEYWORDS)


def is_infutive_related(txn: dict) -> bool:
    return "INFUTIVE" in txn["description"].upper()


def is_interest_income(txn: dict) -> bool:
    desc = txn["description"].upper()
    if txn["deposit"] <= 0:
        return False
    if desc.startswith("SAVING A/C CREDIT INTEREST"):
        return True
    return ("INTEREST" in desc and "CAPITALISED" in desc) or desc == "CREDIT INTEREST CAPITALISED ON SB A/C"


def is_cash_deposit(txn: dict) -> bool:
    desc = txn["description"].upper()
    return txn["deposit"] > 0 and (
        "CASH DEPOSIT" in desc or "BNA CASH" in desc or "CHQ DEPOSIT" in desc
    )


def is_contra_entry(txn: dict) -> bool:
    """Inter-account transfers between YES Bank and Standard Chartered."""
    desc = txn["description"].upper()
    raw = txn["description"]
    bank = txn["bank"]

    if bank == "YES Bank":
        if "STANDARD CHARTERED" in desc:
            return True
        if re.search(r"X{3,4}0545|54410610545", desc):
            return True

    if bank == "Standard Chartered Bank":
        if YES_ACCOUNT in raw:
            if any(
                marker in desc
                for marker in (
                    "ARUN SHANKAR AWASTHI",
                    "ARUNSHANKARAWASTHI",
                    "PAID VIA CRED",
                    "SELF TRANSFER",
                )
            ):
                return True
            if YES_IFSC in desc and "8882283917" in raw:
                return True
            if "IMPS" in desc and "YES BANK" in desc:
                return True

    return False


def get_contra_direction(txn: dict) -> tuple[str, str, str]:
    """Return (direction label, from bank, to bank)."""
    if txn["bank"] == "YES Bank":
        if txn["deposit"] > 0:
            return "SCB → YES Bank", "Standard Chartered Bank", "YES Bank"
        return "YES Bank → SCB", "YES Bank", "Standard Chartered Bank"

    if txn["withdrawal"] > 0:
        return "SCB → YES Bank", "Standard Chartered Bank", "YES Bank"
    return "YES Bank → SCB", "YES Bank", "Standard Chartered Bank"


def contra_amount_by_direction(contra_txns: list[dict]) -> tuple[float, float]:
    scb_to_yes = 0.0
    yes_to_scb = 0.0
    for txn in contra_txns:
        amount = txn["deposit"] or txn["withdrawal"]
        direction, _, _ = get_contra_direction(txn)
        if direction == "SCB → YES Bank":
            scb_to_yes += amount
        else:
            yes_to_scb += amount
    return scb_to_yes, yes_to_scb


def match_contra_pairs(contra_txns: list[dict]) -> list[dict]:
    """Match likely contra pairs by amount within a short date window."""
    outgoings = []
    incomings = []
    for txn in contra_txns:
        direction, from_bank, to_bank = get_contra_direction(txn)
        amount = txn["deposit"] or txn["withdrawal"]
        entry = {**txn, "direction": direction, "from_bank": from_bank, "to_bank": to_bank, "amount": amount}
        if txn["bank"] == from_bank and txn["withdrawal"] > 0:
            outgoings.append(entry)
        elif txn["bank"] == to_bank and txn["deposit"] > 0:
            incomings.append(entry)

    pairs = []
    used_out: set[int] = set()
    used_in: set[int] = set()
    pair_no = 1

    for o_idx, outgoing in enumerate(outgoings):
        if o_idx in used_out or not outgoing["txn_date"]:
            continue
        for i_idx, incoming in enumerate(incomings):
            if i_idx in used_in or not incoming["txn_date"]:
                continue
            if outgoing["direction"] != incoming["direction"]:
                continue
            if abs(outgoing["amount"] - incoming["amount"]) > 0.01:
                continue
            day_gap = abs((outgoing["txn_date"] - incoming["txn_date"]).days)
            if day_gap > 5:
                continue
            pairs.append(
                {
                    "pair_no": pair_no,
                    "amount": outgoing["amount"],
                    "direction": outgoing["direction"],
                    "debit_date": outgoing["txn_date"],
                    "debit_bank": outgoing["bank"],
                    "debit_ref": outgoing["ref_no"],
                    "credit_date": incoming["txn_date"],
                    "credit_bank": incoming["bank"],
                    "credit_ref": incoming["ref_no"],
                    "day_gap": day_gap,
                }
            )
            used_out.add(o_idx)
            used_in.add(i_idx)
            pair_no += 1
            break

    return pairs


def is_self_transfer(txn: dict) -> bool:
    return is_contra_entry(txn)


def classify_income(txn: dict) -> str:
    if is_contra_entry(txn):
        direction, _, _ = get_contra_direction(txn)
        return f"Contra Entry ({direction})"
    if is_infutive_salary(txn):
        return "Salary - Infutive Technology"
    if is_interest_income(txn):
        return "Interest Income (Savings Bank)"
    desc = txn["description"].upper()
    if "PAYTM MONEY" in desc:
        return "Investment Redemption (Paytm Money)"
    if is_cash_deposit(txn):
        return "Cash Deposit"
    if txn["deposit"] > 0 and ("NEFT CR" in desc or "/FROM:" in desc or "NEFT " in desc):
        return "Other Credits (Transfers/Receipts)"
    return "Other"


def style_header_row(ws, row_num: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_table(
    ws,
    start_row: int,
    headers: list[str],
    rows: list[list],
    money_cols: set[int] | None = None,
    date_cols: set[int] | None = None,
) -> int:
    money_cols = money_cols or set()
    date_cols = date_cols or set()

    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    style_header_row(ws, start_row, len(headers))

    for r_offset, row in enumerate(rows, 1):
        r_idx = start_row + r_offset
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx in money_cols and isinstance(value, (int, float)):
                cell.number_format = MONEY_FMT
            if c_idx in date_cols and isinstance(value, datetime):
                cell.number_format = DATE_FMT

    return start_row + len(rows)


def autosize_columns(ws, min_width: int = 10, max_width: int = 55) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def create_itr_summary_sheet(wb: Workbook, transactions: list[dict]) -> None:
    ws = wb.active
    ws.title = "ITR Summary"

    salary_txns = [t for t in transactions if is_infutive_salary(t)]
    interest_txns = [t for t in transactions if is_interest_income(t)]
    cash_txns = [t for t in transactions if is_cash_deposit(t)]
    paytm_redemptions = [
        t for t in transactions if "paytm money" in t["description"].lower() and t["deposit"] > 0
    ]

    yes_txns = [t for t in transactions if t["bank"] == "YES Bank"]
    scb_txns = [t for t in transactions if t["bank"] == "Standard Chartered Bank"]
    contra_txns = [t for t in transactions if is_contra_entry(t)]
    scb_to_yes, yes_to_scb = contra_amount_by_direction(contra_txns)

    row = 1
    ws.cell(row=row, column=1, value="ITR Filing Summary - Combined Bank Statement Analysis").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"{ASSESSEE['financial_year']} | {ASSESSEE['assessment_year']}")
    row += 2

    details = [
        ("Assessee Name", ASSESSEE["name"]),
        ("Email", ASSESSEE["email"]),
        ("Phone", ASSESSEE["phone"]),
        ("Statement Period", ASSESSEE["statement_period"]),
        ("Sources Combined", "YES Bank Excel + Standard Chartered PDF"),
        ("Total Transactions", len(transactions)),
        ("YES Bank Transactions", len(yes_txns)),
        ("Standard Chartered Transactions", len(scb_txns)),
    ]
    for label, value in details:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Bank Accounts Covered").font = TITLE_FONT
    row += 1
    account_rows = [
        ["Bank", "Account No.", "IFSC", "Branch", "Source File"],
        *[
            [a["bank"], a["account_no"], a["ifsc"], a["branch"], a["source"]]
            for a in ACCOUNTS
        ],
    ]
    row = write_table(ws, row, account_rows[0], account_rows[1:]) + 2

    ws.cell(row=row, column=1, value="Income Summary (for ITR)").font = TITLE_FONT
    row += 1
    income_rows = [
        ["Income Head", "Amount (INR)", "No. of Entries", "ITR Schedule / Notes"],
        [
            "Contra Entries (Inter-bank transfers)",
            scb_to_yes + yes_to_scb,
            len(contra_txns),
            "Not taxable - See 'Contra Entries' sheet",
        ],
        [
            "Salary from Infutive Technology Pvt Ltd (Combined)",
            sum(t["deposit"] for t in salary_txns),
            len(salary_txns),
            "Schedule S - See 'Infutive Salary' sheet",
        ],
        [
            "Interest from Savings Bank Accounts (Combined)",
            sum(t["deposit"] for t in interest_txns),
            len(interest_txns),
            "Schedule OS / Section 80TTA (exempt up to Rs.10,000)",
        ],
        [
            "Investment Redemption (Paytm Money)",
            sum(t["deposit"] for t in paytm_redemptions),
            len(paytm_redemptions),
            "Schedule CG / OS - Capital gains or other sources as applicable",
        ],
        [
            "Cash Deposits (Combined)",
            sum(t["deposit"] for t in cash_txns),
            len(cash_txns),
            "Disclose in ITR cash transaction schedule if applicable",
        ],
        [
            "Gross Total Credits (Both Accounts)",
            sum(t["deposit"] for t in transactions),
            sum(1 for t in transactions if t["deposit"] > 0),
            "Not all credits are taxable income",
        ],
    ]
    row = write_table(ws, row, income_rows[0], income_rows[1:], money_cols={2}) + 2

    ws.cell(row=row, column=1, value="Bank-wise Credit / Debit Summary").font = TITLE_FONT
    row += 1
    bank_rows = [["Bank", "Account No.", "Total Deposits", "Total Withdrawals", "Net Flow"]]
    for account in ACCOUNTS:
        bank_txns = [t for t in transactions if t["account_no"] == account["account_no"]]
        deposits = sum(t["deposit"] for t in bank_txns)
        withdrawals = sum(t["withdrawal"] for t in bank_txns)
        bank_rows.append(
            [account["bank"], account["account_no"], deposits, withdrawals, deposits - withdrawals]
        )
    row = write_table(ws, row, bank_rows[0], bank_rows[1:], money_cols={3, 4, 5}) + 2

    ws.cell(row=row, column=1, value="Important Notes for ITR Filing").font = TITLE_FONT
    notes = [
        "1. This workbook combines YES Bank (Excel) and Standard Chartered (PDF) statements for FY 2025-26.",
        "2. Infutive Technology salary credits are identified from both accounts using SALARY / PAY ARUN markers.",
        "3. Refunds, reversals, and credit-card payments from Infutive are excluded from salary.",
        "4. Contra entries (YES Bank ↔ SCB transfers) are separated and excluded from taxable income.",
        "5. Combined savings bank interest is eligible for deduction u/s 80TTA (up to Rs.10,000).",
        "6. Verify total salary against Form 16 / salary slips before filing ITR.",
        "7. Large non-salary credits and cash deposits should be reconciled with source of funds.",
    ]
    for note in notes:
        row += 1
        ws.cell(row=row, column=1, value=note)

    autosize_columns(ws)


def create_infutive_salary_sheet(wb: Workbook, transactions: list[dict]) -> None:
    ws = wb.create_sheet("Infutive Salary")
    salary_txns = sorted(
        [t for t in transactions if is_infutive_salary(t)],
        key=lambda t: (t["txn_date"] or datetime.min, t["bank"]),
    )

    row = 1
    ws.cell(row=row, column=1, value="Salary from Infutive Technology Private Limited").font = TITLE_FONT
    row += 1
    ws.cell(
        row=row,
        column=1,
        value="Combined salary credits from both bank accounts for Schedule S reporting",
    )
    row += 2

    headers = [
        "S.No.",
        "Bank",
        "Account No.",
        "Transaction Date",
        "Value Date",
        "Reference No.",
        "Description",
        "Salary Amount (INR)",
        "Running Balance",
        "Remarks",
    ]
    data = []
    for idx, txn in enumerate(salary_txns, 1):
        data.append(
            [
                idx,
                txn["bank"],
                txn["account_no"],
                txn["txn_date"],
                txn["value_date"],
                txn["ref_no"],
                txn["description"],
                txn["deposit"],
                txn["balance"],
                "Employer salary credit",
            ]
        )

    total = sum(t["deposit"] for t in salary_txns)
    data.append(["", "", "", "", "", "", "TOTAL SALARY (Infutive Technology)", total, "", ""])

    end_row = write_table(ws, row, headers, data, money_cols={8, 9}, date_cols={4, 5})
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=end_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = SUBHEADER_FILL

    row = end_row + 3
    ws.cell(row=row, column=1, value="Other Infutive Technology Entries (Not Salary)").font = TITLE_FONT
    row += 1
    other_headers = ["Bank", "Transaction Date", "Type", "Amount (INR)", "Description", "Suggested Treatment"]
    other_data = []
    for txn in sorted(
        [t for t in transactions if is_infutive_related(t) and not is_infutive_salary(t)],
        key=lambda t: t["txn_date"] or datetime.min,
    ):
        if txn["withdrawal"] > 0:
            other_data.append(
                [
                    txn["bank"],
                    txn["txn_date"],
                    "Debit (Payment to Infutive)",
                    txn["withdrawal"],
                    txn["description"],
                    "Loan repayment / transfer - not salary income",
                ]
            )
        elif txn["deposit"] > 0:
            other_data.append(
                [
                    txn["bank"],
                    txn["txn_date"],
                    "Credit (Excluded from salary)",
                    txn["deposit"],
                    txn["description"],
                    "Refund / reversal / reimbursement - review manually",
                ]
            )

    write_table(ws, row, other_headers, other_data, money_cols={4}, date_cols={2})
    autosize_columns(ws)


def create_interest_sheet(wb: Workbook, transactions: list[dict]) -> None:
    ws = wb.create_sheet("Interest Income")
    interest_txns = sorted(
        [t for t in transactions if is_interest_income(t)],
        key=lambda t: t["txn_date"] or datetime.min,
    )

    headers = ["S.No.", "Bank", "Date", "Description", "Interest Amount (INR)", "Quarter"]
    rows = []
    for idx, txn in enumerate(interest_txns, 1):
        quarter = ""
        if txn["txn_date"]:
            month = txn["txn_date"].month
            if month in (4, 5, 6):
                quarter = "Q1 (Apr-Jun)"
            elif month in (7, 8, 9):
                quarter = "Q2 (Jul-Sep)"
            elif month in (10, 11, 12):
                quarter = "Q3 (Oct-Dec)"
            else:
                quarter = "Q4 (Jan-Mar)"
        rows.append([idx, txn["bank"], txn["txn_date"], txn["description"], txn["deposit"], quarter])

    rows.append(["", "", "", "TOTAL", sum(t["deposit"] for t in interest_txns), "Deduction u/s 80TTA available"])
    write_table(ws, 1, headers, rows, money_cols={5}, date_cols={3})
    autosize_columns(ws)


def create_cash_deposits_sheet(wb: Workbook, transactions: list[dict]) -> None:
    ws = wb.create_sheet("Cash Deposits")
    cash_txns = sorted(
        [t for t in transactions if is_cash_deposit(t)],
        key=lambda t: t["txn_date"] or datetime.min,
    )

    headers = [
        "S.No.",
        "Bank",
        "Transaction Date",
        "Reference No.",
        "Description",
        "Deposit Amount (INR)",
        "Source to Declare in ITR",
    ]
    rows = [
        [
            idx,
            txn["bank"],
            txn["txn_date"],
            txn["ref_no"],
            txn["description"],
            txn["deposit"],
            "Self / Business / Other - verify source",
        ]
        for idx, txn in enumerate(cash_txns, 1)
    ]
    rows.append(["", "", "", "", "TOTAL CASH DEPOSITS", sum(t["deposit"] for t in cash_txns), ""])
    write_table(ws, 1, headers, rows, money_cols={6}, date_cols={3})
    autosize_columns(ws)


def create_contra_entries_sheet(wb: Workbook, transactions: list[dict]) -> None:
    ws = wb.create_sheet("Contra Entries")
    contra_txns = sorted(
        [t for t in transactions if is_contra_entry(t)],
        key=lambda t: t["txn_date"] or datetime.min,
    )
    pairs = match_contra_pairs(contra_txns)
    scb_to_yes, yes_to_scb = contra_amount_by_direction(contra_txns)

    row = 1
    ws.cell(row=row, column=1, value="Contra Entries - Inter-Bank Transfers (Own Accounts)").font = TITLE_FONT
    row += 1
    ws.cell(
        row=row,
        column=1,
        value="Transfers between YES Bank (014091900000507) and Standard Chartered (54410610545) - not taxable income",
    )
    row += 2

    summary_rows = [
        ["Summary", "Amount (INR)", "Count", "Notes"],
        ["SCB → YES Bank (money moved to YES Bank)", scb_to_yes, sum(1 for t in contra_txns if get_contra_direction(t)[0] == "SCB → YES Bank"), "Debit in SCB or Credit in YES"],
        ["YES Bank → SCB (money moved to SCB)", yes_to_scb, sum(1 for t in contra_txns if get_contra_direction(t)[0] == "YES Bank → SCB"), "Debit in YES or Credit in SCB"],
        ["Total Contra Volume", scb_to_yes + yes_to_scb, len(contra_txns), "Both directions combined"],
        ["Matched Pairs Identified", sum(p["amount"] for p in pairs), len(pairs), "Same amount within 5 days"],
    ]
    row = write_table(ws, row, summary_rows[0], summary_rows[1:], money_cols={2}) + 2

    ws.cell(row=row, column=1, value="All Contra Entries").font = TITLE_FONT
    row += 1
    headers = [
        "S.No.",
        "Direction",
        "From Bank",
        "To Bank",
        "Transaction Date",
        "Value Date",
        "Type",
        "Amount (INR)",
        "Bank (Entry Side)",
        "Account No.",
        "Reference No.",
        "Description",
    ]
    data = []
    for idx, txn in enumerate(contra_txns, 1):
        direction, from_bank, to_bank = get_contra_direction(txn)
        amount = txn["deposit"] or txn["withdrawal"]
        txn_type = "Credit" if txn["deposit"] > 0 else "Debit"
        data.append(
            [
                idx,
                direction,
                from_bank,
                to_bank,
                txn["txn_date"],
                txn["value_date"],
                txn_type,
                amount,
                txn["bank"],
                txn["account_no"],
                txn["ref_no"],
                txn["description"],
            ]
        )

    row = write_table(ws, row, headers, data, money_cols={8}, date_cols={5, 6}) + 2

    ws.cell(row=row, column=1, value="Matched Contra Pairs (Same Amount, Within 5 Days)").font = TITLE_FONT
    row += 1
    pair_headers = [
        "Pair No.",
        "Direction",
        "Amount (INR)",
        "Debit Date",
        "Debit Bank",
        "Debit Ref",
        "Credit Date",
        "Credit Bank",
        "Credit Ref",
        "Day Gap",
    ]
    pair_rows = [
        [
            pair["pair_no"],
            pair["direction"],
            pair["amount"],
            pair["debit_date"],
            pair["debit_bank"],
            pair["debit_ref"],
            pair["credit_date"],
            pair["credit_bank"],
            pair["credit_ref"],
            pair["day_gap"],
        ]
        for pair in pairs
    ]
    if not pair_rows:
        pair_rows = [["", "", "", "", "", "", "", "", "", "No auto-matched pairs found"]]

    write_table(ws, row, pair_headers, pair_rows, money_cols={3}, date_cols={4, 7})
    autosize_columns(ws)


def create_monthly_summary_sheet(wb: Workbook, transactions: list[dict]) -> None:
    ws = wb.create_sheet("Monthly Summary")
    monthly: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"deposits": 0.0, "withdrawals": 0.0, "count": 0}
    )

    for txn in transactions:
        if not txn["txn_date"]:
            continue
        key = (txn["bank"], txn["txn_date"].strftime("%b-%Y"))
        monthly[key]["deposits"] += txn["deposit"]
        monthly[key]["withdrawals"] += txn["withdrawal"]
        monthly[key]["count"] += 1

    headers = ["Bank", "Month", "No. of Txns", "Total Deposits (INR)", "Total Withdrawals (INR)", "Net (INR)"]
    rows = []
    for (bank, month) in sorted(monthly.keys(), key=lambda item: (item[0], datetime.strptime(item[1], "%b-%Y"))):
        data = monthly[(bank, month)]
        rows.append([bank, month, data["count"], data["deposits"], data["withdrawals"], data["deposits"] - data["withdrawals"]])

    write_table(ws, 1, headers, rows, money_cols={4, 5, 6})
    autosize_columns(ws)


def create_all_transactions_sheet(wb: Workbook, transactions: list[dict]) -> None:
    ws = wb.create_sheet("All Transactions")
    headers = [
        "S.No.",
        "Bank",
        "Account No.",
        "Transaction Date",
        "Value Date",
        "Reference No.",
        "Description",
        "Withdrawal (INR)",
        "Deposit (INR)",
        "Balance (INR)",
        "Category",
    ]

    sorted_txns = sorted(transactions, key=lambda t: t["txn_date"] or datetime.min, reverse=True)
    rows = []
    for idx, txn in enumerate(sorted_txns, 1):
        category = classify_income(txn) if txn["deposit"] > 0 else "Expense/Transfer"
        rows.append(
            [
                idx,
                txn["bank"],
                txn["account_no"],
                txn["txn_date"],
                txn["value_date"],
                txn["ref_no"],
                txn["description"],
                txn["withdrawal"] or None,
                txn["deposit"] or None,
                txn["balance"],
                category,
            ]
        )

    write_table(ws, 1, headers, rows, money_cols={8, 9, 10}, date_cols={4, 5})
    autosize_columns(ws)


def create_bank_wise_sheets(wb: Workbook, transactions: list[dict]) -> None:
    for account in ACCOUNTS:
        bank_txns = [t for t in transactions if t["account_no"] == account["account_no"]]
        sheet_name = "YES Bank Txns" if account["bank"] == "YES Bank" else "SCB Txns"
        ws = wb.create_sheet(sheet_name)
        headers = [
            "S.No.",
            "Transaction Date",
            "Value Date",
            "Reference No.",
            "Description",
            "Withdrawal (INR)",
            "Deposit (INR)",
            "Balance (INR)",
        ]
        rows = [
            [
                idx,
                txn["txn_date"],
                txn["value_date"],
                txn["ref_no"],
                txn["description"],
                txn["withdrawal"] or None,
                txn["deposit"] or None,
                txn["balance"],
            ]
            for idx, txn in enumerate(
                sorted(bank_txns, key=lambda t: t["txn_date"] or datetime.min, reverse=True),
                1,
            )
        ]
        write_table(ws, 1, headers, rows, money_cols={6, 7, 8}, date_cols={2, 3})
        autosize_columns(ws)


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    transactions = load_all_transactions()

    wb = Workbook()
    create_itr_summary_sheet(wb, transactions)
    create_infutive_salary_sheet(wb, transactions)
    create_contra_entries_sheet(wb, transactions)
    create_interest_sheet(wb, transactions)
    create_cash_deposits_sheet(wb, transactions)
    create_monthly_summary_sheet(wb, transactions)
    create_bank_wise_sheets(wb, transactions)
    create_all_transactions_sheet(wb, transactions)

    wb.save(OUTPUT_PATH)

    salary_txns = [t for t in transactions if is_infutive_salary(t)]
    contra_txns = [t for t in transactions if is_contra_entry(t)]
    print(f"Created: {OUTPUT_PATH}")
    print(f"YES Bank transactions: {len([t for t in transactions if t['bank'] == 'YES Bank'])}")
    print(f"SCB transactions: {len([t for t in transactions if t['bank'] == 'Standard Chartered Bank'])}")
    print(f"Combined transactions: {len(transactions)}")
    print(f"Contra entries: {len(contra_txns)}")
    print(f"Infutive salary entries: {len(salary_txns)}")
    print(f"Total salary amount: {sum(t['deposit'] for t in salary_txns):,.2f}")
    print(f"YES Bank salary: {sum(t['deposit'] for t in salary_txns if t['bank'] == 'YES Bank'):,.2f}")
    print(f"SCB salary: {sum(t['deposit'] for t in salary_txns if t['bank'] == 'Standard Chartered Bank'):,.2f}")


if __name__ == "__main__":
    main()
